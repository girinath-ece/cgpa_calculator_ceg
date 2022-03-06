"""
For Windows
"""

from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
import sys
import os

from openpyxl.cell import cell
from openpyxl import Workbook

class NewUser(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.title = 'New User Registration'
        self.initUi()

    def initUi(self):
        self.setWindowTitle(self.title)
        self.setFixedSize(500,300)

        self.name_label = QtWidgets.QLabel('Full Name:')
        self.roll_label = QtWidgets.QLabel('Roll Number:')
        self.phno_label = QtWidgets.QLabel('Contact Number:')
        self.mail_label = QtWidgets.QLabel('E-Mail ID:')
        self.dept_label = QtWidgets.QLabel('Department:')

        self.name_edit = QtWidgets.QLineEdit()      
        self.roll_edit = QtWidgets.QLineEdit()
        self.phno_edit = QtWidgets.QLineEdit()
        self.mail_edit = QtWidgets.QLineEdit()
         
        self.dept_edit = QtWidgets.QComboBox(self)
        department = [" ",
                      "CSE",
                      "IT",
                      "ECE",
                      "BIO MEDICAL",
                      "EEE",
                      "PRINTING",
                      "MINING",
                      "MANUFACTURING",
                      "INDUSTRIAL",
                      "MECHANICAL",
                      "MATERIAL SCIENCE",
                      "CIVIL",
                      "GEO INFORMATICS"]
        
        self.dept_edit.addItems(department)
        
        self.grid = QtWidgets.QGridLayout()
        self.grid.setSpacing(10)
        self.grid.addWidget(self.name_label,0,0)
        self.grid.addWidget(self.roll_label,1,0)
        self.grid.addWidget(self.phno_label,2,0)
        self.grid.addWidget(self.mail_label,3,0)
        self.grid.addWidget(self.dept_label,4,0)

        self.grid.addWidget(self.name_edit,0,1)
        self.grid.addWidget(self.roll_edit,1,1)
        self.grid.addWidget(self.phno_edit,2,1)
        self.grid.addWidget(self.mail_edit,3,1)
        self.grid.addWidget(self.dept_edit,4,1)

        self.save_button = QtWidgets.QPushButton(self)
        self.cancel_button = QtWidgets.QPushButton(self)
        self.clear_button = QtWidgets.QPushButton(self)
        self.save_button.setText("Save")
        self.cancel_button.setText("Cancel")
        self.clear_button.setText("Clear")

        self.save_button.clicked.connect(self.save_data)
        self.clear_button.clicked.connect(self.clear_data)
        self.cancel_button.clicked.connect(self.cancel_data)

        self.hbox = QtWidgets.QHBoxLayout()
        self.hbox.addStretch(1)

        self.hbox.addWidget(self.clear_button) 
        self.hbox.addWidget(self.cancel_button) 
        self.hbox.addWidget(self.save_button)
        
        self.vbox = QtWidgets.QVBoxLayout()
        self.vbox.addLayout(self.grid)
        self.vbox.addLayout(self.hbox)

        self.setLayout(self.vbox)
    
    def save_data(self):
        print('save')
        name = self.name_edit.text()
        print(name)
        roll = self.roll_edit.text()
        print(roll)
        phone = self.phno_edit.text()
        print(phone)
        mail = self.mail_edit.text()
        print(mail)
        department = self.dept_edit.currentText()
        print(len(department))
        usr_name = os.getlogin()
        wb = openpyxl.load_workbook("gpa.xlsx")
        sheet = wb.active
        flag = 0

        # check whether the roll number is already registered
        for i in range(2,sheet.max_row+1):
            cell_obj = sheet.cell(row=i,column=2)
            if(roll == cell_obj.value):
                self.show_roll_no_warning()
                flag = 1
        
        # check whether all the details have been filled
        # check whether the ph num and roll num is of length 10
        if (name == '' or roll == '' or phone == '' or mail =='' or department==' ' or len(roll)!=10 or len(phone)!=10 or not(roll.isdigit()) or not(phone.isdigit())):
            flag = 2
            self.show_null_warning()

        # check whether valid email id has been entered
        mail_domains = ['gmail.com','yahoo.com','annauniv.edu.in','outlook.com','hotmail.com','icloud.com']
        mail_dom = mail.split('@')[-1]
        if mail_dom not in mail_domains:
            flag = 3
            self.show_invalid_mail()

        if flag ==0:
            sheet.append((name,roll,phone,mail,department))
            self.show_success()
        wb.save("gpa.xlsx")

    def clear_data(self):
        print('clear')
        self.name_edit.clear()
        self.roll_edit.clear()
        self.phno_edit.clear()
        self.mail_edit.clear()
        self.dept_edit.setCurrentIndex(0)

    def cancel_data(self):
        print('cancel')
        self.close()

    def show_roll_no_warning(self):
        print('roll number already exists') 
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        msg.setWindowTitle('Warning')
        msg.setText('Roll Number is already registered.')
        msg.setInformativeText('Kindly Update!')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()
    
    def show_null_warning(self):
        print('some data provided is null') 
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        msg.setWindowTitle('Warning')
        msg.setText('Check and fill all the details.')
        msg.setInformativeText('Roll number and phone number must be a valid 10 digit number.')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()
     
    def show_invalid_mail(self):
        print('mail is incorrect') 
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        msg.setWindowTitle('Warning')
        msg.setText('Domain of your mail ID is invalid.')
        msg.setInformativeText('Try using \'gmail.com\',\'yahoo.com\',\'annauniv.edu.in\',\'outlook.com\',\'hotmail.com\',\'icloud.com\' ')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()

    def show_success(self):
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setWindowTitle('Registered')
        msg.setText('Registration Successful!')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()

class UpdateUser(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.title = 'Update User Information'
        self.initUi()
    
    def initUi(self):
        self.setWindowTitle(self.title)
        self.setFixedSize(500,600)
        self.name_label = QtWidgets.QLabel('Full Name:')
        self.roll_label = QtWidgets.QLabel('Roll Number:')
        self.phno_label = QtWidgets.QLabel('Contact Number:')
        self.mail_label = QtWidgets.QLabel('E-Mail ID:')
        self.dept_label = QtWidgets.QLabel('Department:')
        
        self.name_edit = QtWidgets.QLineEdit()   
        self.roll_edit = QtWidgets.QLineEdit()
        self.phno_edit = QtWidgets.QLineEdit() 
        self.mail_edit = QtWidgets.QLineEdit()

        self.dept_edit = QtWidgets.QComboBox(self)
        department = [" ",
                      "CSE",
                      "IT",
                      "ECE",
                      "BIO MEDICAL",
                      "EEE",
                      "PRINTING",
                      "MINING",
                      "MANUFACTURING",
                      "INDUSTRIAL",
                      "MECHANICAL",
                      "MATERIAL SCIENCE",
                      "CIVIL",
                      "GEO INFORMATICS"]
        
        self.dept_edit.addItems(department)
        
        self.grid = QtWidgets.QGridLayout()
        self.grid.setSpacing(10)
        self.grid.addWidget(self.name_label,0,0)
        self.grid.addWidget(self.roll_label,1,0)
        self.grid.addWidget(self.phno_label,2,0)
        self.grid.addWidget(self.mail_label,3,0)
        self.grid.addWidget(self.dept_label,4,0)
        
        self.grid.addWidget(self.name_edit,0,1)
        self.grid.addWidget(self.roll_edit,1,1)
        self.grid.addWidget(self.phno_edit,2,1)
        self.grid.addWidget(self.mail_edit,3,1)
        self.grid.addWidget(self.dept_edit,4,1)

        ################################################

        self.tabs = QtWidgets.QTabWidget()
        self.sem1 = QtWidgets.QWidget()
        self.sem2 = QtWidgets.QWidget()
        self.sem3 = QtWidgets.QWidget()
        self.sem4 = QtWidgets.QWidget()
        self.sem5 = QtWidgets.QWidget()
        self.sem6 = QtWidgets.QWidget()
        self.sem7 = QtWidgets.QWidget()
        self.sem8 = QtWidgets.QWidget()
        
        self.tabs.addTab(self.sem1,"Semester 1")
        self.tabs.addTab(self.sem2,"Semester 2")
        self.tabs.addTab(self.sem3,"Semester 3")
        self.tabs.addTab(self.sem4,"Semester 4")
        self.tabs.addTab(self.sem5,"Semester 5")
        self.tabs.addTab(self.sem6,"Semester 6")
        self.tabs.addTab(self.sem7,"Semester 7")
        self.tabs.addTab(self.sem8,"Semester 8")

        # Semester 1
        self.sem1.layout = QtWidgets.QGridLayout()
        self.sem1.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem1_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem1_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub1_name_label = QtWidgets.QLabel('English                        ')
        self.sem1_sub1_grade = QtWidgets.QLineEdit()

        self.sem1_sub2_credit_label = QtWidgets.QLabel('4')
        self.sem1_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub2_name_label = QtWidgets.QLabel('Maths 1')
        self.sem1_sub2_grade = QtWidgets.QLineEdit()

        self.sem1_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem1_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub3_name_label = QtWidgets.QLabel('Physics')
        self.sem1_sub3_grade = QtWidgets.QLineEdit()

        self.sem1_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem1_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub4_name_label = QtWidgets.QLabel('Chemistry')
        self.sem1_sub4_grade = QtWidgets.QLineEdit()

        self.sem1_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem1_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub5_name_label = QtWidgets.QLabel('Python')
        self.sem1_sub5_grade = QtWidgets.QLineEdit()
        
        self.sem1_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem1_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub6_name_label = QtWidgets.QLabel('Science Lab')
        self.sem1_sub6_grade = QtWidgets.QLineEdit()

        self.sem1_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem1_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub7_name_label = QtWidgets.QLabel('Python Lab')
        self.sem1_sub7_grade = QtWidgets.QLineEdit()

        self.sem1.layout.addWidget(self.title_label1,0,0)
        self.sem1.layout.addWidget(self.title_label2,0,1)
        self.sem1.layout.addWidget(self.title_label3,0,2)

        self.sem1.layout.addWidget(self.sem1_sub1_credit_label,1,0)
        self.sem1.layout.addWidget(self.sem1_sub1_name_label,1,1)
        self.sem1.layout.addWidget(self.sem1_sub1_grade,1,2)
        
        self.sem1.layout.addWidget(self.sem1_sub2_credit_label,2,0)
        self.sem1.layout.addWidget(self.sem1_sub2_name_label,2,1)
        self.sem1.layout.addWidget(self.sem1_sub2_grade,2,2)
        
        self.sem1.layout.addWidget(self.sem1_sub3_credit_label,3,0)
        self.sem1.layout.addWidget(self.sem1_sub3_name_label,3,1)
        self.sem1.layout.addWidget(self.sem1_sub3_grade,3,2)

        self.sem1.layout.addWidget(self.sem1_sub4_credit_label,4,0)
        self.sem1.layout.addWidget(self.sem1_sub4_name_label,4,1)
        self.sem1.layout.addWidget(self.sem1_sub4_grade,4,2)

        self.sem1.layout.addWidget(self.sem1_sub5_credit_label,5,0)
        self.sem1.layout.addWidget(self.sem1_sub5_name_label,5,1)
        self.sem1.layout.addWidget(self.sem1_sub5_grade,5,2)

        self.sem1.layout.addWidget(self.sem1_sub6_credit_label,6,0)
        self.sem1.layout.addWidget(self.sem1_sub6_name_label,6,1)
        self.sem1.layout.addWidget(self.sem1_sub6_grade,6,2)

        self.sem1.layout.addWidget(self.sem1_sub7_credit_label,7,0)
        self.sem1.layout.addWidget(self.sem1_sub7_name_label,7,1)
        self.sem1.layout.addWidget(self.sem1_sub7_grade,7,2)

        self.sem1.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem1.setLayout(self.sem1.layout)

        # Semester 2
        self.sem2.layout = QtWidgets.QGridLayout()
        self.sem2.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem2_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem2_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub1_name_label = QtWidgets.QLabel('Maths 2                        ')
        self.sem2_sub1_grade = QtWidgets.QLineEdit()

        self.sem2_sub2_credit_label = QtWidgets.QLabel('4')
        self.sem2_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub2_name_label = QtWidgets.QLabel('Mechanics')
        self.sem2_sub2_grade = QtWidgets.QLineEdit()

        self.sem2_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem2_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub3_name_label = QtWidgets.QLabel('Circuit Theory')
        self.sem2_sub3_grade = QtWidgets.QLineEdit()

        self.sem2_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem2_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub4_name_label = QtWidgets.QLabel('Electrical & Measurement')
        self.sem2_sub4_grade = QtWidgets.QLineEdit()

        self.sem2_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem2_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub5_name_label = QtWidgets.QLabel('Semiconductors')
        self.sem2_sub5_grade = QtWidgets.QLineEdit()
        
        self.sem2_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem2_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub6_name_label = QtWidgets.QLabel('Circuit Theory Lab')
        self.sem2_sub6_grade = QtWidgets.QLineEdit()

        self.sem2_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem2_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub7_name_label = QtWidgets.QLabel('Workshop Lab')
        self.sem2_sub7_grade = QtWidgets.QLineEdit()

        self.sem2.layout.addWidget(self.title_label1,0,0)
        self.sem2.layout.addWidget(self.title_label2,0,1)
        self.sem2.layout.addWidget(self.title_label3,0,2)

        self.sem2.layout.addWidget(self.sem2_sub1_credit_label,1,0)
        self.sem2.layout.addWidget(self.sem2_sub1_name_label,1,1)
        self.sem2.layout.addWidget(self.sem2_sub1_grade,1,2)
        
        self.sem2.layout.addWidget(self.sem2_sub2_credit_label,2,0)
        self.sem2.layout.addWidget(self.sem2_sub2_name_label,2,1)
        self.sem2.layout.addWidget(self.sem2_sub2_grade,2,2)
        
        self.sem2.layout.addWidget(self.sem2_sub3_credit_label,3,0)
        self.sem2.layout.addWidget(self.sem2_sub3_name_label,3,1)
        self.sem2.layout.addWidget(self.sem2_sub3_grade,3,2)

        self.sem2.layout.addWidget(self.sem2_sub4_credit_label,4,0)
        self.sem2.layout.addWidget(self.sem2_sub4_name_label,4,1)
        self.sem2.layout.addWidget(self.sem2_sub4_grade,4,2)

        self.sem2.layout.addWidget(self.sem2_sub5_credit_label,5,0)
        self.sem2.layout.addWidget(self.sem2_sub5_name_label,5,1)
        self.sem2.layout.addWidget(self.sem2_sub5_grade,5,2)

        self.sem2.layout.addWidget(self.sem2_sub6_credit_label,6,0)
        self.sem2.layout.addWidget(self.sem2_sub6_name_label,6,1)
        self.sem2.layout.addWidget(self.sem2_sub6_grade,6,2)

        self.sem2.layout.addWidget(self.sem2_sub7_credit_label,7,0)
        self.sem2.layout.addWidget(self.sem2_sub7_name_label,7,1)
        self.sem2.layout.addWidget(self.sem2_sub7_grade,7,2)

        self.sem2.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem2.setLayout(self.sem2.layout)


        # Semester 3
        self.sem3.layout = QtWidgets.QGridLayout()
        self.sem3.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem3_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem3_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub1_name_label = QtWidgets.QLabel('Linear Algebra')
        self.sem3_sub1_grade = QtWidgets.QLineEdit()

        self.sem3_sub2_credit_label = QtWidgets.QLabel('4')
        self.sem3_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub2_name_label = QtWidgets.QLabel('Electronic Circuits I')
        self.sem3_sub2_grade = QtWidgets.QLineEdit()

        self.sem3_sub3_credit_label = QtWidgets.QLabel('4')
        self.sem3_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub3_name_label = QtWidgets.QLabel('Signals & System')
        self.sem3_sub3_grade = QtWidgets.QLineEdit()

        self.sem3_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem3_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub4_name_label = QtWidgets.QLabel('Electromagnetic Fields & Waves ')
        self.sem3_sub4_grade = QtWidgets.QLineEdit()

        self.sem3_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem3_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub5_name_label = QtWidgets.QLabel('Digital System Design')
        self.sem3_sub5_grade = QtWidgets.QLineEdit()
        
        self.sem3_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem3_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub6_name_label = QtWidgets.QLabel('Electronic Design Lab')
        self.sem3_sub6_grade = QtWidgets.QLineEdit()

        self.sem3_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem3_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub7_name_label = QtWidgets.QLabel('Electrical & Measurements Lab')
        self.sem3_sub7_grade = QtWidgets.QLineEdit()

        self.sem3.layout.addWidget(self.title_label1,0,0)
        self.sem3.layout.addWidget(self.title_label2,0,1)
        self.sem3.layout.addWidget(self.title_label3,0,2)

        self.sem3.layout.addWidget(self.sem3_sub1_credit_label,1,0)
        self.sem3.layout.addWidget(self.sem3_sub1_name_label,1,1)
        self.sem3.layout.addWidget(self.sem3_sub1_grade,1,2)
        
        self.sem3.layout.addWidget(self.sem3_sub2_credit_label,2,0)
        self.sem3.layout.addWidget(self.sem3_sub2_name_label,2,1)
        self.sem3.layout.addWidget(self.sem3_sub2_grade,2,2)
        
        self.sem3.layout.addWidget(self.sem3_sub3_credit_label,3,0)
        self.sem3.layout.addWidget(self.sem3_sub3_name_label,3,1)
        self.sem3.layout.addWidget(self.sem3_sub3_grade,3,2)

        self.sem3.layout.addWidget(self.sem3_sub4_credit_label,4,0)
        self.sem3.layout.addWidget(self.sem3_sub4_name_label,4,1)
        self.sem3.layout.addWidget(self.sem3_sub4_grade,4,2)

        self.sem3.layout.addWidget(self.sem3_sub5_credit_label,5,0)
        self.sem3.layout.addWidget(self.sem3_sub5_name_label,5,1)
        self.sem3.layout.addWidget(self.sem3_sub5_grade,5,2)

        self.sem3.layout.addWidget(self.sem3_sub6_credit_label,6,0)
        self.sem3.layout.addWidget(self.sem3_sub6_name_label,6,1)
        self.sem3.layout.addWidget(self.sem3_sub6_grade,6,2)

        self.sem3.layout.addWidget(self.sem3_sub7_credit_label,7,0)
        self.sem3.layout.addWidget(self.sem3_sub7_name_label,7,1)
        self.sem3.layout.addWidget(self.sem3_sub7_grade,7,2)

        self.sem3.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem3.setLayout(self.sem3.layout)

     
        # Semester 4
        self.sem4.layout = QtWidgets.QGridLayout()
        self.sem4.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem4_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem4_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub1_name_label = QtWidgets.QLabel('Electronic Circuits II')
        self.sem4_sub1_grade = QtWidgets.QLineEdit()

        self.sem4_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub2_name_label = QtWidgets.QLabel('Transmission Line & Wave Guides')
        self.sem4_sub2_grade = QtWidgets.QLineEdit()

        self.sem4_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub3_name_label = QtWidgets.QLabel('Communication Theory')
        self.sem4_sub3_grade = QtWidgets.QLineEdit()

        self.sem4_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub4_name_label = QtWidgets.QLabel('Digital Signal Processing')
        self.sem4_sub4_grade = QtWidgets.QLineEdit()

        self.sem4_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub5_name_label = QtWidgets.QLabel('Linear Intergrated Circuits')
        self.sem4_sub5_grade = QtWidgets.QLineEdit()
       
        self.sem4_sub6_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub6_name_label = QtWidgets.QLabel('Environmental Science')
        self.sem4_sub6_grade = QtWidgets.QLineEdit()

        self.sem4_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem4_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub7_name_label = QtWidgets.QLabel('Digital Signal Lab')
        self.sem4_sub7_grade = QtWidgets.QLineEdit()

        self.sem4_sub8_credit_label = QtWidgets.QLabel('2')
        self.sem4_sub8_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub8_name_label = QtWidgets.QLabel('Integrated Circuits Lab')
        self.sem4_sub8_grade = QtWidgets.QLineEdit()

        self.sem4.layout.addWidget(self.title_label1,0,0)
        self.sem4.layout.addWidget(self.title_label2,0,1)
        self.sem4.layout.addWidget(self.title_label3,0,2)

        self.sem4.layout.addWidget(self.sem4_sub1_credit_label,1,0)
        self.sem4.layout.addWidget(self.sem4_sub1_name_label,1,1)
        self.sem4.layout.addWidget(self.sem4_sub1_grade,1,2)
        
        self.sem4.layout.addWidget(self.sem4_sub2_credit_label,2,0)
        self.sem4.layout.addWidget(self.sem4_sub2_name_label,2,1)
        self.sem4.layout.addWidget(self.sem4_sub2_grade,2,2)
        
        self.sem4.layout.addWidget(self.sem4_sub3_credit_label,3,0)
        self.sem4.layout.addWidget(self.sem4_sub3_name_label,3,1)
        self.sem4.layout.addWidget(self.sem4_sub3_grade,3,2)

        self.sem4.layout.addWidget(self.sem4_sub4_credit_label,4,0)
        self.sem4.layout.addWidget(self.sem4_sub4_name_label,4,1)
        self.sem4.layout.addWidget(self.sem4_sub4_grade,4,2)

        self.sem4.layout.addWidget(self.sem4_sub5_credit_label,5,0)
        self.sem4.layout.addWidget(self.sem4_sub5_name_label,5,1)
        self.sem4.layout.addWidget(self.sem4_sub5_grade,5,2)

        self.sem4.layout.addWidget(self.sem4_sub6_credit_label,6,0)
        self.sem4.layout.addWidget(self.sem4_sub6_name_label,6,1)
        self.sem4.layout.addWidget(self.sem4_sub6_grade,6,2)

        self.sem4.layout.addWidget(self.sem4_sub7_credit_label,7,0)
        self.sem4.layout.addWidget(self.sem4_sub7_name_label,7,1)
        self.sem4.layout.addWidget(self.sem4_sub7_grade,7,2)

        self.sem4.layout.addWidget(self.sem4_sub8_credit_label,8,0)
        self.sem4.layout.addWidget(self.sem4_sub8_name_label,8,1)
        self.sem4.layout.addWidget(self.sem4_sub8_grade,8,2)

        self.sem4.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem4.setLayout(self.sem4.layout)

        # Semester 5
        self.sem5.layout = QtWidgets.QGridLayout()
        self.sem5.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem5_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub1_name_label = QtWidgets.QLabel('Antennas')
        self.sem5_sub1_grade = QtWidgets.QLineEdit()

        self.sem5_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub2_name_label = QtWidgets.QLabel('Digital Communication')
        self.sem5_sub2_grade = QtWidgets.QLineEdit()

        self.sem5_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub3_name_label = QtWidgets.QLabel('Microprocessors')
        self.sem5_sub3_grade = QtWidgets.QLineEdit()

        self.sem5_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub4_name_label = QtWidgets.QLabel('Control Systems')
        self.sem5_sub4_grade = QtWidgets.QLineEdit()

        self.sem5_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub5_name_label = QtWidgets.QLabel('Principles of Management       ')
        self.sem5_sub5_grade = QtWidgets.QLineEdit()
        
        self.sem5_sub6_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub6_name_label = QtWidgets.QLabel('Professinal Elective 1')
        self.sem5_sub6_grade = QtWidgets.QLineEdit()

        self.sem5_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem5_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub7_name_label = QtWidgets.QLabel('Microprocessors Lab')
        self.sem5_sub7_grade = QtWidgets.QLineEdit()

        self.sem5_sub8_credit_label = QtWidgets.QLabel('2')
        self.sem5_sub8_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub8_name_label = QtWidgets.QLabel('Digital Communication Lab')
        self.sem5_sub8_grade = QtWidgets.QLineEdit()

        self.sem5_sub9_credit_label = QtWidgets.QLabel('2')
        self.sem5_sub9_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub9_name_label = QtWidgets.QLabel('Summer Project')
        self.sem5_sub9_grade = QtWidgets.QLineEdit()

        self.sem5.layout.addWidget(self.title_label1,0,0)
        self.sem5.layout.addWidget(self.title_label2,0,1)
        self.sem5.layout.addWidget(self.title_label3,0,2)

        self.sem5.layout.addWidget(self.sem5_sub1_credit_label,1,0)
        self.sem5.layout.addWidget(self.sem5_sub1_name_label,1,1)
        self.sem5.layout.addWidget(self.sem5_sub1_grade,1,2)

        self.sem5.layout.addWidget(self.sem5_sub2_credit_label,2,0)
        self.sem5.layout.addWidget(self.sem5_sub2_name_label,2,1)
        self.sem5.layout.addWidget(self.sem5_sub2_grade,2,2)

        self.sem5.layout.addWidget(self.sem5_sub3_credit_label,3,0)
        self.sem5.layout.addWidget(self.sem5_sub3_name_label,3,1)
        self.sem5.layout.addWidget(self.sem5_sub3_grade,3,2)

        self.sem5.layout.addWidget(self.sem5_sub4_credit_label,4,0)
        self.sem5.layout.addWidget(self.sem5_sub4_name_label,4,1)
        self.sem5.layout.addWidget(self.sem5_sub4_grade,4,2)

        self.sem5.layout.addWidget(self.sem5_sub5_credit_label,5,0)
        self.sem5.layout.addWidget(self.sem5_sub5_name_label,5,1)
        self.sem5.layout.addWidget(self.sem5_sub5_grade,5,2)

        self.sem5.layout.addWidget(self.sem5_sub6_credit_label,6,0)
        self.sem5.layout.addWidget(self.sem5_sub6_name_label,6,1)
        self.sem5.layout.addWidget(self.sem5_sub6_grade,6,2)

        self.sem5.layout.addWidget(self.sem5_sub7_credit_label,7,0)
        self.sem5.layout.addWidget(self.sem5_sub7_name_label,7,1)
        self.sem5.layout.addWidget(self.sem5_sub7_grade,7,2)

        self.sem5.layout.addWidget(self.sem5_sub8_credit_label,8,0)
        self.sem5.layout.addWidget(self.sem5_sub8_name_label,8,1)
        self.sem5.layout.addWidget(self.sem5_sub8_grade,8,2)

        self.sem5.layout.addWidget(self.sem5_sub9_credit_label,9,0)
        self.sem5.layout.addWidget(self.sem5_sub9_name_label,9,1)
        self.sem5.layout.addWidget(self.sem5_sub9_grade,9,2)

        self.sem5.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem5.setLayout(self.sem5.layout)

        # Semester 6
        self.sem6.layout = QtWidgets.QGridLayout()
        self.sem6.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem6_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub1_name_label = QtWidgets.QLabel('Digital VLSI')
        self.sem6_sub1_grade = QtWidgets.QLineEdit()

        self.sem6_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub2_name_label = QtWidgets.QLabel('Wireless Communications        ')
        self.sem6_sub2_grade = QtWidgets.QLineEdit()

        self.sem6_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub3_name_label = QtWidgets.QLabel('Communication Networks')
        self.sem6_sub3_grade = QtWidgets.QLineEdit()

        self.sem6_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub4_name_label = QtWidgets.QLabel('Professional Elective II')
        self.sem6_sub4_grade = QtWidgets.QLineEdit()

        self.sem6_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub5_name_label = QtWidgets.QLabel('Open Elective I')
        self.sem6_sub5_grade = QtWidgets.QLineEdit()
        
        self.sem6_sub6_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub6_name_label = QtWidgets.QLabel('Values & Ethics')
        self.sem6_sub6_grade = QtWidgets.QLineEdit()

        self.sem6_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem6_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub7_name_label = QtWidgets.QLabel('VLSI Lab')
        self.sem6_sub7_grade = QtWidgets.QLineEdit()

        self.sem6_sub8_credit_label = QtWidgets.QLabel('2')
        self.sem6_sub8_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub8_name_label = QtWidgets.QLabel('Wireless Communications Lab')
        self.sem6_sub8_grade = QtWidgets.QLineEdit()

        self.sem6.layout.addWidget(self.title_label1,0,0)
        self.sem6.layout.addWidget(self.title_label2,0,1)
        self.sem6.layout.addWidget(self.title_label3,0,2)

        self.sem6.layout.addWidget(self.sem6_sub1_credit_label,1,0)
        self.sem6.layout.addWidget(self.sem6_sub1_name_label,1,1)
        self.sem6.layout.addWidget(self.sem6_sub1_grade,1,2)
        
        self.sem6.layout.addWidget(self.sem6_sub2_credit_label,2,0)
        self.sem6.layout.addWidget(self.sem6_sub2_name_label,2,1)
        self.sem6.layout.addWidget(self.sem6_sub2_grade,2,2)
        
        self.sem6.layout.addWidget(self.sem6_sub3_credit_label,3,0)
        self.sem6.layout.addWidget(self.sem6_sub3_name_label,3,1)
        self.sem6.layout.addWidget(self.sem6_sub3_grade,3,2)

        self.sem6.layout.addWidget(self.sem6_sub4_credit_label,4,0)
        self.sem6.layout.addWidget(self.sem6_sub4_name_label,4,1)
        self.sem6.layout.addWidget(self.sem6_sub4_grade,4,2)

        self.sem6.layout.addWidget(self.sem6_sub5_credit_label,5,0)
        self.sem6.layout.addWidget(self.sem6_sub5_name_label,5,1)
        self.sem6.layout.addWidget(self.sem6_sub5_grade,5,2)

        self.sem6.layout.addWidget(self.sem6_sub6_credit_label,6,0)
        self.sem6.layout.addWidget(self.sem6_sub6_name_label,6,1)
        self.sem6.layout.addWidget(self.sem6_sub6_grade,6,2)

        self.sem6.layout.addWidget(self.sem6_sub7_credit_label,7,0)
        self.sem6.layout.addWidget(self.sem6_sub7_name_label,7,1)
        self.sem6.layout.addWidget(self.sem6_sub7_grade,7,2)

        self.sem6.layout.addWidget(self.sem6_sub8_credit_label,8,0)
        self.sem6.layout.addWidget(self.sem6_sub8_name_label,8,1)
        self.sem6.layout.addWidget(self.sem6_sub8_grade,8,2)

        self.sem6.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem6.setLayout(self.sem6.layout)

        # Semester 7
        self.sem7.layout = QtWidgets.QGridLayout()
        self.sem7.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem7_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub1_name_label = QtWidgets.QLabel('Optical Wave Communications    ')
        self.sem7_sub1_grade = QtWidgets.QLineEdit()

        self.sem7_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub2_name_label = QtWidgets.QLabel('Human Relations')
        self.sem7_sub2_grade = QtWidgets.QLineEdit()

        self.sem7_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub3_name_label = QtWidgets.QLabel('Professional Elective III')
        self.sem7_sub3_grade = QtWidgets.QLineEdit()

        self.sem7_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub4_name_label = QtWidgets.QLabel('Professional Elective IV')
        self.sem7_sub4_grade = QtWidgets.QLineEdit()

        self.sem7_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub5_name_label = QtWidgets.QLabel('Open Elective II')
        self.sem7_sub5_grade = QtWidgets.QLineEdit()
        
        self.sem7_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem7_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub6_name_label = QtWidgets.QLabel('High Frequency Lab')
        self.sem7_sub6_grade = QtWidgets.QLineEdit()

        self.sem7_sub7_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub7_name_label = QtWidgets.QLabel('Project I')
        self.sem7_sub7_grade = QtWidgets.QLineEdit()

        self.sem7.layout.addWidget(self.title_label1,0,0)
        self.sem7.layout.addWidget(self.title_label2,0,1)
        self.sem7.layout.addWidget(self.title_label3,0,2)

        self.sem7.layout.addWidget(self.sem7_sub1_credit_label,1,0)
        self.sem7.layout.addWidget(self.sem7_sub1_name_label,1,1)
        self.sem7.layout.addWidget(self.sem7_sub1_grade,1,2)
        
        self.sem7.layout.addWidget(self.sem7_sub2_credit_label,2,0)
        self.sem7.layout.addWidget(self.sem7_sub2_name_label,2,1)
        self.sem7.layout.addWidget(self.sem7_sub2_grade,2,2)
        
        self.sem7.layout.addWidget(self.sem7_sub3_credit_label,3,0)
        self.sem7.layout.addWidget(self.sem7_sub3_name_label,3,1)
        self.sem7.layout.addWidget(self.sem7_sub3_grade,3,2)

        self.sem7.layout.addWidget(self.sem7_sub4_credit_label,4,0)
        self.sem7.layout.addWidget(self.sem7_sub4_name_label,4,1)
        self.sem7.layout.addWidget(self.sem7_sub4_grade,4,2)

        self.sem7.layout.addWidget(self.sem7_sub5_credit_label,5,0)
        self.sem7.layout.addWidget(self.sem7_sub5_name_label,5,1)
        self.sem7.layout.addWidget(self.sem7_sub5_grade,5,2)

        self.sem7.layout.addWidget(self.sem7_sub6_credit_label,6,0)
        self.sem7.layout.addWidget(self.sem7_sub6_name_label,6,1)
        self.sem7.layout.addWidget(self.sem7_sub6_grade,6,2)

        self.sem7.layout.addWidget(self.sem7_sub7_credit_label,7,0)
        self.sem7.layout.addWidget(self.sem7_sub7_name_label,7,1)
        self.sem7.layout.addWidget(self.sem7_sub7_grade,7,2)

        self.sem7.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem7.setLayout(self.sem7.layout)

        # Semester 8
        self.sem8.layout = QtWidgets.QGridLayout()
        self.sem8.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem8_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem8_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem8_sub1_name_label = QtWidgets.QLabel('Professional Elective V')
        self.sem8_sub1_grade = QtWidgets.QLineEdit()

        self.sem8_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem8_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem8_sub2_name_label = QtWidgets.QLabel('Professional Elective VI       ')
        self.sem8_sub2_grade = QtWidgets.QLineEdit()

        self.sem8_sub3_credit_label = QtWidgets.QLabel('8')
        self.sem8_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem8_sub3_name_label = QtWidgets.QLabel('Project II')
        self.sem8_sub3_grade = QtWidgets.QLineEdit()
        
        self.sem8.layout.addWidget(self.title_label1,0,0)
        self.sem8.layout.addWidget(self.title_label2,0,1)
        self.sem8.layout.addWidget(self.title_label3,0,2)

        self.sem8.layout.addWidget(self.sem8_sub1_credit_label,1,0)
        self.sem8.layout.addWidget(self.sem8_sub1_name_label,1,1)
        self.sem8.layout.addWidget(self.sem8_sub1_grade,1,2)
        
        self.sem8.layout.addWidget(self.sem8_sub2_credit_label,2,0)
        self.sem8.layout.addWidget(self.sem8_sub2_name_label,2,1)
        self.sem8.layout.addWidget(self.sem8_sub2_grade,2,2)
        
        self.sem8.layout.addWidget(self.sem8_sub3_credit_label,3,0)
        self.sem8.layout.addWidget(self.sem8_sub3_name_label,3,1)
        self.sem8.layout.addWidget(self.sem8_sub3_grade,3,2)
        
        self.sem8.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem8.setLayout(self.sem8.layout)

        self.hbox_tab = QtWidgets.QHBoxLayout()
        self.hbox_tab.addWidget(self.tabs)

        ######################################################
        self.save_button = QtWidgets.QPushButton(self)
        self.cancel_button = QtWidgets.QPushButton(self)
        self.clear_button = QtWidgets.QPushButton(self)
        self.search_button = QtWidgets.QPushButton(self)

        self.save_button.setText("Save")
        self.cancel_button.setText("Cancel")
        self.clear_button.setText("Clear")
        self.search_button.setText("Search")

        self.save_button.clicked.connect(self.save_data)
        self.clear_button.clicked.connect(self.clear_data)
        self.cancel_button.clicked.connect(self.cancel_data)
        self.search_button.clicked.connect(self.search_data)

        self.hbox = QtWidgets.QHBoxLayout()
        self.hbox.addStretch(1)

        self.hbox.addWidget(self.search_button)
        self.hbox.addWidget(self.clear_button) 
        self.hbox.addWidget(self.cancel_button) 
        self.hbox.addWidget(self.save_button)
    
        ######################################################


        self.vbox = QtWidgets.QVBoxLayout()
        self.vbox.addLayout(self.grid)
        self.vbox.addLayout(self.hbox_tab)
        self.vbox.addLayout(self.hbox)

        self.setLayout(self.vbox)
        self.disable_line_edit(0)
           
    

    def search_data(self):
        print('searching ',self.roll_edit.text())
        roll = self.roll_edit.text()
        department=[" ",
                    "CSE",
                    "IT",
                    "ECE",
                    "BIO MEDICAL",
                    "EEE",
                    "PRINTING",
                    "MINING",
                    "MANUFACTURING",
                    "INDUSTRIAL",
                    "MECHANICAL",
                    "MATERIAL SCIENCE",
                    "CIVIL",
                    "GEO INFORMATICS"]


        wb = openpyxl.load_workbook("gpa.xlsx")
        sheet = wb.active
        available_flag = 0
        for i in range(2,sheet.max_row+1):
            cell_obj = sheet.cell(row=i,column=2)
            if roll == cell_obj.value:
                available_flag = 1

        if available_flag == 1:
            for i in range(2,sheet.max_row+1):
                cell_obj = sheet.cell(row=i,column=2)
                if(roll == cell_obj.value):
                    self.name_edit.setText((sheet.cell(row=i,column=1)).value)
                    self.phno_edit.setText((sheet.cell(row=i,column=3)).value)
                    self.mail_edit.setText((sheet.cell(row=i,column=4)).value)
                    self.dept_edit.setCurrentIndex(department.index((sheet.cell(row=i,column=5)).value))
                    self.sem1_sub1_grade.setText((sheet.cell(row=i,column=6)).value)
                    self.sem1_sub2_grade.setText((sheet.cell(row=i,column=7)).value)
                    self.sem1_sub3_grade.setText((sheet.cell(row=i,column=8)).value)
                    self.sem1_sub4_grade.setText((sheet.cell(row=i,column=9)).value)
                    self.sem1_sub5_grade.setText((sheet.cell(row=i,column=10)).value)
                    self.sem1_sub6_grade.setText((sheet.cell(row=i,column=11)).value)
                    self.sem1_sub7_grade.setText((sheet.cell(row=i,column=12)).value)
                    self.sem2_sub1_grade.setText((sheet.cell(row=i,column=13)).value)
                    self.sem2_sub2_grade.setText((sheet.cell(row=i,column=14)).value)
                    self.sem2_sub3_grade.setText((sheet.cell(row=i,column=15)).value)
                    self.sem2_sub4_grade.setText((sheet.cell(row=i,column=16)).value)
                    self.sem2_sub5_grade.setText((sheet.cell(row=i,column=17)).value)
                    self.sem2_sub6_grade.setText((sheet.cell(row=i,column=18)).value)
                    self.sem2_sub7_grade.setText((sheet.cell(row=i,column=19)).value)
                    self.sem3_sub1_grade.setText((sheet.cell(row=i,column=20)).value)
                    self.sem3_sub2_grade.setText((sheet.cell(row=i,column=21)).value)
                    self.sem3_sub3_grade.setText((sheet.cell(row=i,column=22)).value)
                    self.sem3_sub4_grade.setText((sheet.cell(row=i,column=23)).value)
                    self.sem3_sub5_grade.setText((sheet.cell(row=i,column=24)).value)
                    self.sem3_sub6_grade.setText((sheet.cell(row=i,column=25)).value)
                    self.sem3_sub7_grade.setText((sheet.cell(row=i,column=26)).value)
                    self.sem4_sub1_grade.setText((sheet.cell(row=i,column=27)).value)
                    self.sem4_sub2_grade.setText((sheet.cell(row=i,column=28)).value)
                    self.sem4_sub3_grade.setText((sheet.cell(row=i,column=29)).value)
                    self.sem4_sub4_grade.setText((sheet.cell(row=i,column=30)).value)
                    self.sem4_sub5_grade.setText((sheet.cell(row=i,column=31)).value)
                    self.sem4_sub6_grade.setText((sheet.cell(row=i,column=32)).value)
                    self.sem4_sub7_grade.setText((sheet.cell(row=i,column=33)).value)
                    self.sem4_sub8_grade.setText((sheet.cell(row=i,column=34)).value)
                    self.sem5_sub1_grade.setText((sheet.cell(row=i,column=35)).value)
                    self.sem5_sub2_grade.setText((sheet.cell(row=i,column=36)).value)
                    self.sem5_sub3_grade.setText((sheet.cell(row=i,column=37)).value)
                    self.sem5_sub4_grade.setText((sheet.cell(row=i,column=38)).value)
                    self.sem5_sub5_grade.setText((sheet.cell(row=i,column=39)).value)
                    self.sem5_sub6_grade.setText((sheet.cell(row=i,column=40)).value)
                    self.sem5_sub7_grade.setText((sheet.cell(row=i,column=41)).value)
                    self.sem5_sub8_grade.setText((sheet.cell(row=i,column=42)).value)
                    self.sem5_sub9_grade.setText((sheet.cell(row=i,column=43)).value)
                    self.sem6_sub1_grade.setText((sheet.cell(row=i,column=44)).value)
                    self.sem6_sub2_grade.setText((sheet.cell(row=i,column=45)).value)
                    self.sem6_sub3_grade.setText((sheet.cell(row=i,column=46)).value)
                    self.sem6_sub4_grade.setText((sheet.cell(row=i,column=47)).value)
                    self.sem6_sub5_grade.setText((sheet.cell(row=i,column=48)).value)
                    self.sem6_sub6_grade.setText((sheet.cell(row=i,column=49)).value)
                    self.sem6_sub7_grade.setText((sheet.cell(row=i,column=50)).value)
                    self.sem6_sub8_grade.setText((sheet.cell(row=i,column=51)).value)
                    self.sem7_sub1_grade.setText((sheet.cell(row=i,column=52)).value)
                    self.sem7_sub2_grade.setText((sheet.cell(row=i,column=53)).value)
                    self.sem7_sub3_grade.setText((sheet.cell(row=i,column=54)).value)
                    self.sem7_sub4_grade.setText((sheet.cell(row=i,column=55)).value)
                    self.sem7_sub5_grade.setText((sheet.cell(row=i,column=56)).value)
                    self.sem7_sub6_grade.setText((sheet.cell(row=i,column=57)).value)
                    self.sem7_sub7_grade.setText((sheet.cell(row=i,column=58)).value)
                    self.sem8_sub1_grade.setText((sheet.cell(row=i,column=59)).value)
                    self.sem8_sub2_grade.setText((sheet.cell(row=i,column=60)).value)
                    self.sem8_sub3_grade.setText((sheet.cell(row=i,column=61)).value)
                    
                    self.disable_line_edit(1)
                    break
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setWindowTitle('Warning')
            msg.setText('Roll Number is Invalid!')
            msg.setInformativeText('Kindly register or enter a valid roll number.')
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            retval = msg.exec_()




    def disable_line_edit(self,state):
        self.name_edit.setEnabled(bool(state))
        self.roll_edit.setEnabled(bool(1-state))
        self.phno_edit.setEnabled(bool(state))
        self.mail_edit.setEnabled(bool(state))
        self.dept_edit.setEnabled(bool(state))
        self.sem1_sub1_grade.setEnabled(bool(state))
        self.sem1_sub2_grade.setEnabled(bool(state))
        self.sem1_sub3_grade.setEnabled(bool(state))
        self.sem1_sub4_grade.setEnabled(bool(state))
        self.sem1_sub5_grade.setEnabled(bool(state))
        self.sem1_sub6_grade.setEnabled(bool(state))
        self.sem1_sub7_grade.setEnabled(bool(state))

        self.sem2_sub1_grade.setEnabled(bool(state))
        self.sem2_sub2_grade.setEnabled(bool(state))
        self.sem2_sub3_grade.setEnabled(bool(state))
        self.sem2_sub4_grade.setEnabled(bool(state))
        self.sem2_sub5_grade.setEnabled(bool(state))
        self.sem2_sub6_grade.setEnabled(bool(state))
        self.sem2_sub7_grade.setEnabled(bool(state))

        self.sem3_sub1_grade.setEnabled(bool(state))
        self.sem3_sub2_grade.setEnabled(bool(state))
        self.sem3_sub3_grade.setEnabled(bool(state))
        self.sem3_sub4_grade.setEnabled(bool(state))
        self.sem3_sub5_grade.setEnabled(bool(state))
        self.sem3_sub6_grade.setEnabled(bool(state))
        self.sem3_sub7_grade.setEnabled(bool(state))

        self.sem4_sub1_grade.setEnabled(bool(state))
        self.sem4_sub2_grade.setEnabled(bool(state))
        self.sem4_sub3_grade.setEnabled(bool(state))
        self.sem4_sub4_grade.setEnabled(bool(state))
        self.sem4_sub5_grade.setEnabled(bool(state))
        self.sem4_sub6_grade.setEnabled(bool(state))
        self.sem4_sub7_grade.setEnabled(bool(state))
        self.sem4_sub8_grade.setEnabled(bool(state))

        self.sem5_sub1_grade.setEnabled(bool(state))
        self.sem5_sub2_grade.setEnabled(bool(state))
        self.sem5_sub3_grade.setEnabled(bool(state))
        self.sem5_sub4_grade.setEnabled(bool(state))
        self.sem5_sub5_grade.setEnabled(bool(state))
        self.sem5_sub6_grade.setEnabled(bool(state))
        self.sem5_sub7_grade.setEnabled(bool(state))
        self.sem5_sub8_grade.setEnabled(bool(state))
        self.sem5_sub9_grade.setEnabled(bool(state))

        self.sem6_sub1_grade.setEnabled(bool(state))
        self.sem6_sub2_grade.setEnabled(bool(state))
        self.sem6_sub3_grade.setEnabled(bool(state))
        self.sem6_sub4_grade.setEnabled(bool(state))
        self.sem6_sub5_grade.setEnabled(bool(state))
        self.sem6_sub6_grade.setEnabled(bool(state))
        self.sem6_sub7_grade.setEnabled(bool(state))
        self.sem6_sub8_grade.setEnabled(bool(state))

        self.sem7_sub1_grade.setEnabled(bool(state))
        self.sem7_sub2_grade.setEnabled(bool(state))
        self.sem7_sub3_grade.setEnabled(bool(state))
        self.sem7_sub4_grade.setEnabled(bool(state))
        self.sem7_sub5_grade.setEnabled(bool(state))
        self.sem7_sub6_grade.setEnabled(bool(state))
        self.sem7_sub7_grade.setEnabled(bool(state))

        self.sem8_sub1_grade.setEnabled(bool(state))
        self.sem8_sub2_grade.setEnabled(bool(state))
        self.sem8_sub3_grade.setEnabled(bool(state))

                  
    def save_data(self):
        print('save')
        name = self.name_edit.text()
        print(name)
        roll = self.roll_edit.text()
        print(roll)
        phone = self.phno_edit.text()
        print(phone)
        mail = self.mail_edit.text()
        print(mail)
        department = self.dept_edit.currentText()
        print(department)

        sem1_grade = [self.sem1_sub1_grade.text(),
                      self.sem1_sub2_grade.text(),
                      self.sem1_sub3_grade.text(),
                      self.sem1_sub4_grade.text(),
                      self.sem1_sub5_grade.text(),
                      self.sem1_sub6_grade.text(),
                      self.sem1_sub7_grade.text()]

        sem2_grade = [self.sem2_sub1_grade.text(),
                      self.sem2_sub2_grade.text(),
                      self.sem2_sub3_grade.text(),
                      self.sem2_sub4_grade.text(),
                      self.sem2_sub5_grade.text(),
                      self.sem2_sub6_grade.text(),
                      self.sem2_sub7_grade.text()]
        
        sem3_grade = [self.sem3_sub1_grade.text(),
                      self.sem3_sub2_grade.text(),
                      self.sem3_sub3_grade.text(),
                      self.sem3_sub4_grade.text(),
                      self.sem3_sub5_grade.text(),
                      self.sem3_sub6_grade.text(),
                      self.sem3_sub7_grade.text()]

        sem4_grade = [self.sem4_sub1_grade.text(),
                      self.sem4_sub2_grade.text(),
                      self.sem4_sub3_grade.text(),
                      self.sem4_sub4_grade.text(),
                      self.sem4_sub5_grade.text(),
                      self.sem4_sub6_grade.text(),
                      self.sem4_sub7_grade.text(),
                      self.sem4_sub8_grade.text()]
        
        sem5_grade = [self.sem5_sub1_grade.text(),
                      self.sem5_sub2_grade.text(),
                      self.sem5_sub3_grade.text(),
                      self.sem5_sub4_grade.text(),
                      self.sem5_sub5_grade.text(),
                      self.sem5_sub6_grade.text(),
                      self.sem5_sub7_grade.text(),
                      self.sem5_sub8_grade.text(),
                      self.sem5_sub9_grade.text()]
                      
        sem6_grade = [self.sem6_sub1_grade.text(),
                      self.sem6_sub2_grade.text(),
                      self.sem6_sub3_grade.text(),
                      self.sem6_sub4_grade.text(),
                      self.sem6_sub5_grade.text(),
                      self.sem6_sub6_grade.text(),
                      self.sem6_sub7_grade.text(),
                      self.sem6_sub8_grade.text()]
        
        sem7_grade = [self.sem7_sub1_grade.text(),
                      self.sem7_sub2_grade.text(),
                      self.sem7_sub3_grade.text(),
                      self.sem7_sub4_grade.text(),
                      self.sem7_sub5_grade.text(),
                      self.sem7_sub6_grade.text(),
                      self.sem7_sub7_grade.text()]
        
        sem8_grade = [self.sem8_sub1_grade.text(),
                      self.sem8_sub2_grade.text(),
                      self.sem8_sub3_grade.text()]
        
        grade_dict = {'sem1':sem1_grade,
                      'sem2':sem2_grade,
                      'sem3':sem3_grade,
                      'sem4':sem4_grade,
                      'sem5':sem5_grade,
                      'sem6':sem6_grade,
                      'sem7':sem7_grade,
                      'sem8':sem8_grade}


        
        grade_flag = 0

        ref_mark = {'O':10,'A+':9,'A':8,'B+':7,'B':6,'C+':5,'RA':0,'WH':0,'None':0,'':0}
        
        # check whether grades are valid
        for grade_key,grade_value in grade_dict.items():
            for grade in grade_value:
                if grade not in ref_mark.keys():
                    grade_flag = -1
                    self.show_invalid_grade()
                    

        if grade_flag==0:
            # mark_dict
            mark_dict = {}
            for grade_key,grade_value in grade_dict.items():
                mark_val = []
                for grade in grade_value:
                    mk = ref_mark[grade]
                    mark_val.append(mk)
                mark_dict[grade_key] = mark_val
            

            # to calculate gpa
            credit_score = {'sem1':[4.0,4.0,3.0,3.0,3.0,2.0,2.0], 
                            'sem2':[4.0,4.0,4.0,3.0,3.0,2.0,2.0], 
                            'sem3':[4.0,4.0,4.0,3.0,3.0,2.0,2.0], 
                            'sem4':[4.0,3.0,3.0,3.0,3.0,3.0,2.0,2.0], 
                            'sem5':[3.0,3.0,3.0,3.0,3.0,3.0,2.0,2.0,2.0], 
                            'sem6':[3.0,3.0,3.0,3.0,3.0,3.0,2.0,2.0], 
                            'sem7':[3.0,3.0,3.0,3.0,3.0,2.0,3.0], 
                            'sem8':[3.0,3.0,8.0] }
            gpa = {}
            cgpa_sum = 0.0
            total_credit = 0.0
            for mark_key,mark_value in mark_dict.items():
                gpa_sum = 0.00
                crd_score = credit_score[mark_key]
                if(sum(mark_value)!=0):
                    for crd,mk in zip(crd_score,mark_value):
                        gpa_sum +=crd*mk
                    
                    cgpa_sum+=gpa_sum
                    total_credit += sum(crd_score)
                    gpa[mark_key] = round(gpa_sum/(sum(crd_score)),2)

            cgpa = cgpa_sum/total_credit
            cgpa = round(cgpa,2)
            print('cgpa : ',cgpa)
            print(gpa)
            print(grade_dict)
            print(mark_dict)

        usr_name = os.getlogin()
        wb = openpyxl.load_workbook("gpa.xlsx")
        sheet = wb.active
        flag = 0

        # check whether all the details have been filled
        # check whether the ph num and roll num is of length 10
        if (name == '' or phone == '' or mail =='' or department==' ' or len(phone)!=10 or not(phone.isdigit())):
            flag = -1
            self.show_null_warning()

        # check whether valid email id has been entered
        mail_domains = ['gmail.com','yahoo.com','annauniv.edu.in','outlook.com','hotmail.com','icloud.com']
        mail_dom = mail.split('@')[-1]
        if mail_dom not in mail_domains:
            flag = -1
            self.show_invalid_mail()

        if flag ==0 and grade_flag == 0:
            for i in range(2,sheet.max_row+1):
                cell_obj = sheet.cell(row=i,column=2)
                if(roll == cell_obj.value):
                    col_ind = 6
                    sheet.cell(row=i,column=1).value = self.name_edit.text()
                    sheet.cell(row=i,column=2).value = self.roll_edit.text()
                    sheet.cell(row=i,column=3).value = self.phno_edit.text()
                    sheet.cell(row=i,column=4).value = self.mail_edit.text()
                    sheet.cell(row=i,column=5).value = self.dept_edit.currentText()
                    
                    for _,val in grade_dict.items():
                        for sub_ind in range(0,len(val)):
                            sheet.cell(row=i,column=col_ind).value = val[sub_ind]
                            col_ind+=1
                    
                    temp_ind = col_ind
                    for _ in range(0,8):
                        sheet.cell(row=i,column=temp_ind).value = 0.00
                        temp_ind+=1
                    
                    
                    for _,gpa_val in gpa.items():
                        sheet.cell(row=i,column=col_ind).value = gpa_val
                        col_ind+=1
                    
                    sheet.cell(row=i,column=70).value = cgpa
            self.show_success()
        
        wb.save("gpa.xlsx")

    def show_invalid_grade(self):
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        msg.setWindowTitle('Warning')
        msg.setText('Enter a valid grade.')
        msg.setInformativeText('The grade can be any of the following: \n \'O\',\'A+\',\'A\',\'B+\',\'B\',\'C+\',\'WH\',\'RA\' ')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()
    
    def show_null_warning(self):
        print('some data provided is null') 
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        msg.setWindowTitle('Warning')
        msg.setText('Check and fill all the details.')
        msg.setInformativeText('Roll number and phone number must be a valid 10 digit number.')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()
     
    def show_invalid_mail(self):
        print('mail is incorrect') 
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        msg.setWindowTitle('Warning')
        msg.setText('Domain of your mail ID is invalid.')
        msg.setInformativeText('Try using \'gmail.com\',\'yahoo.com\',\'annauniv.edu.in\',\'outlook.com\',\'hotmail.com\',\'icloud.com\' ')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()

    def show_success(self):
        msg = QtWidgets.QMessageBox()
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setWindowTitle('Updated')
        msg.setText('Details updated Successfully!')
        msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
        retval = msg.exec_()

    def clear_data(self):
        print('clear')

        self.disable_line_edit(0)
        self.name_edit.clear()
        self.roll_edit.clear()
        self.phno_edit.clear()
        self.mail_edit.clear()
        self.dept_edit.setCurrentIndex(0)

        self.sem1_sub1_grade.clear()
        self.sem1_sub2_grade.clear()
        self.sem1_sub3_grade.clear()
        self.sem1_sub4_grade.clear()
        self.sem1_sub5_grade.clear()
        self.sem1_sub6_grade.clear()
        self.sem1_sub7_grade.clear()

        self.sem2_sub1_grade.clear()
        self.sem2_sub2_grade.clear()
        self.sem2_sub3_grade.clear()
        self.sem2_sub4_grade.clear()
        self.sem2_sub5_grade.clear()
        self.sem2_sub6_grade.clear()
        self.sem2_sub7_grade.clear()

        self.sem3_sub1_grade.clear()
        self.sem3_sub2_grade.clear()
        self.sem3_sub3_grade.clear()
        self.sem3_sub4_grade.clear()
        self.sem3_sub5_grade.clear()
        self.sem3_sub6_grade.clear()
        self.sem3_sub7_grade.clear()

        self.sem4_sub1_grade.clear()
        self.sem4_sub2_grade.clear()
        self.sem4_sub3_grade.clear()
        self.sem4_sub4_grade.clear()
        self.sem4_sub5_grade.clear()
        self.sem4_sub6_grade.clear()
        self.sem4_sub7_grade.clear()
        self.sem4_sub8_grade.clear()

        self.sem5_sub1_grade.clear()
        self.sem5_sub2_grade.clear()
        self.sem5_sub3_grade.clear()
        self.sem5_sub4_grade.clear()
        self.sem5_sub5_grade.clear()
        self.sem5_sub6_grade.clear()
        self.sem5_sub7_grade.clear()
        self.sem5_sub8_grade.clear()
        self.sem5_sub9_grade.clear()

        self.sem6_sub1_grade.clear()
        self.sem6_sub2_grade.clear()
        self.sem6_sub3_grade.clear()
        self.sem6_sub4_grade.clear()
        self.sem6_sub5_grade.clear()
        self.sem6_sub6_grade.clear()
        self.sem6_sub7_grade.clear()
        self.sem6_sub8_grade.clear()

        self.sem7_sub1_grade.clear()
        self.sem7_sub2_grade.clear()
        self.sem7_sub3_grade.clear()
        self.sem7_sub4_grade.clear()
        self.sem7_sub5_grade.clear()
        self.sem7_sub6_grade.clear()
        self.sem7_sub7_grade.clear()

        self.sem8_sub1_grade.clear()
        self.sem8_sub2_grade.clear()
        self.sem8_sub3_grade.clear()
        

    def cancel_data(self):
        print('cancel')
        self.close()

class ViewUser(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.title = 'View User Information'
        self.initUi()
    
    def initUi(self):
        self.setWindowTitle(self.title)
        self.setFixedSize(500,600)
        self.name_label = QtWidgets.QLabel('Full Name:')
        self.roll_label = QtWidgets.QLabel('Roll Number:')
        self.phno_label = QtWidgets.QLabel('Contact Number:')
        self.mail_label = QtWidgets.QLabel('E-Mail ID:')
        self.dept_label = QtWidgets.QLabel('Department:')
        self.cgpa_label = QtWidgets.QLabel('CGPA:')
        
        self.name_edit = QtWidgets.QLabel()      
        self.roll_edit = QtWidgets.QLineEdit()
        self.phno_edit = QtWidgets.QLabel()
        self.mail_edit = QtWidgets.QLabel()
        self.cgpa_edit = QtWidgets.QLabel()
        self.dept_edit = QtWidgets.QLabel()

        department = [" ",
                      "CSE",
                      "IT",
                      "ECE",
                      "BIO MEDICAL",
                      "EEE",
                      "PRINTING",
                      "MINING",
                      "MANUFACTURING",
                      "INDUSTRIAL",
                      "MECHANICAL",
                      "MATERIAL SCIENCE",
                      "CIVIL",
                      "GEO INFORMATICS"]
        
        
        self.grid = QtWidgets.QGridLayout()
        self.grid.setSpacing(10)
        self.grid.addWidget(self.name_label,0,0)
        self.grid.addWidget(self.roll_label,1,0)
        self.grid.addWidget(self.phno_label,2,0)
        self.grid.addWidget(self.mail_label,3,0)
        self.grid.addWidget(self.dept_label,4,0)
        self.grid.addWidget(self.cgpa_label,5,0)
        
        self.grid.addWidget(self.name_edit,0,1)
        self.grid.addWidget(self.roll_edit,1,1)
        self.grid.addWidget(self.phno_edit,2,1)
        self.grid.addWidget(self.mail_edit,3,1)
        self.grid.addWidget(self.dept_edit,4,1)
        self.grid.addWidget(self.cgpa_edit,5,1)

        ################################################


        self.tabs = QtWidgets.QTabWidget()
        self.sem1 = QtWidgets.QWidget()
        self.sem2 = QtWidgets.QWidget()
        self.sem3 = QtWidgets.QWidget()
        self.sem4 = QtWidgets.QWidget()
        self.sem5 = QtWidgets.QWidget()
        self.sem6 = QtWidgets.QWidget()
        self.sem7 = QtWidgets.QWidget()
        self.sem8 = QtWidgets.QWidget()
        
        self.tabs.addTab(self.sem1,"Semester 1")
        self.tabs.addTab(self.sem2,"Semester 2")
        self.tabs.addTab(self.sem3,"Semester 3")
        self.tabs.addTab(self.sem4,"Semester 4")
        self.tabs.addTab(self.sem5,"Semester 5")
        self.tabs.addTab(self.sem6,"Semester 6")
        self.tabs.addTab(self.sem7,"Semester 7")
        self.tabs.addTab(self.sem8,"Semester 8")

        # Semester 1
        self.sem1.layout = QtWidgets.QGridLayout()
        self.sem1.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem1_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem1_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub1_name_label = QtWidgets.QLabel('English')
        self.sem1_sub1_grade_label = QtWidgets.QLabel()

        self.sem1_sub2_credit_label = QtWidgets.QLabel('4')
        self.sem1_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub2_name_label = QtWidgets.QLabel('Maths 1')
        self.sem1_sub2_grade_label = QtWidgets.QLabel()

        self.sem1_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem1_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub3_name_label = QtWidgets.QLabel('Physics')
        self.sem1_sub3_grade_label = QtWidgets.QLabel()

        self.sem1_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem1_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub4_name_label = QtWidgets.QLabel('Chemistry')
        self.sem1_sub4_grade_label = QtWidgets.QLabel()

        self.sem1_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem1_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub5_name_label = QtWidgets.QLabel('Python')
        self.sem1_sub5_grade_label = QtWidgets.QLabel()
        
        self.sem1_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem1_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub6_name_label = QtWidgets.QLabel('Science Lab')
        self.sem1_sub6_grade_label = QtWidgets.QLabel()

        self.sem1_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem1_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_sub7_name_label = QtWidgets.QLabel('Python Lab')
        self.sem1_sub7_grade_label = QtWidgets.QLabel()

        self.sem1.layout.addWidget(self.title_label1,0,0)
        self.sem1.layout.addWidget(self.title_label2,0,1)
        self.sem1.layout.addWidget(self.title_label3,0,2)

        self.sem1.layout.addWidget(self.sem1_sub1_credit_label,1,0)
        self.sem1.layout.addWidget(self.sem1_sub1_name_label,1,1)
        self.sem1.layout.addWidget(self.sem1_sub1_grade_label,1,2)
        
        self.sem1.layout.addWidget(self.sem1_sub2_credit_label,2,0)
        self.sem1.layout.addWidget(self.sem1_sub2_name_label,2,1)
        self.sem1.layout.addWidget(self.sem1_sub2_grade_label,2,2)
        
        self.sem1.layout.addWidget(self.sem1_sub3_credit_label,3,0)
        self.sem1.layout.addWidget(self.sem1_sub3_name_label,3,1)
        self.sem1.layout.addWidget(self.sem1_sub3_grade_label,3,2)

        self.sem1.layout.addWidget(self.sem1_sub4_credit_label,4,0)
        self.sem1.layout.addWidget(self.sem1_sub4_name_label,4,1)
        self.sem1.layout.addWidget(self.sem1_sub4_grade_label,4,2)

        self.sem1.layout.addWidget(self.sem1_sub5_credit_label,5,0)
        self.sem1.layout.addWidget(self.sem1_sub5_name_label,5,1)
        self.sem1.layout.addWidget(self.sem1_sub5_grade_label,5,2)

        self.sem1.layout.addWidget(self.sem1_sub6_credit_label,6,0)
        self.sem1.layout.addWidget(self.sem1_sub6_name_label,6,1)
        self.sem1.layout.addWidget(self.sem1_sub6_grade_label,6,2)

        self.sem1.layout.addWidget(self.sem1_sub7_credit_label,7,0)
        self.sem1.layout.addWidget(self.sem1_sub7_name_label,7,1)
        self.sem1.layout.addWidget(self.sem1_sub7_grade_label,7,2)

        self.sem1_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem1_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem1_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem1.layout.addWidget(self.sem1_gpa_label,8,0)
        self.sem1.layout.addWidget(self.sem1_gpa_result_label,8,2)

        self.sem1.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem1.setLayout(self.sem1.layout)

        # Semester 2
        self.sem2.layout = QtWidgets.QGridLayout()
        self.sem2.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem2_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem2_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub1_name_label = QtWidgets.QLabel('Maths 2                        ')
        self.sem2_sub1_grade_label = QtWidgets.QLabel()

        self.sem2_sub2_credit_label = QtWidgets.QLabel('4')
        self.sem2_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub2_name_label = QtWidgets.QLabel('Mechanics')
        self.sem2_sub2_grade_label = QtWidgets.QLabel()

        self.sem2_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem2_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub3_name_label = QtWidgets.QLabel('Circuit Theory')
        self.sem2_sub3_grade_label = QtWidgets.QLabel()

        self.sem2_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem2_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub4_name_label = QtWidgets.QLabel('Electrical & Measurement')
        self.sem2_sub4_grade_label = QtWidgets.QLabel()

        self.sem2_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem2_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub5_name_label = QtWidgets.QLabel('Semiconductors')
        self.sem2_sub5_grade_label = QtWidgets.QLabel()
        
        self.sem2_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem2_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub6_name_label = QtWidgets.QLabel('Circuit Theory Lab')
        self.sem2_sub6_grade_label = QtWidgets.QLabel()

        self.sem2_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem2_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2_sub7_name_label = QtWidgets.QLabel('Workshop Lab')
        self.sem2_sub7_grade_label = QtWidgets.QLabel()

        self.sem2.layout.addWidget(self.title_label1,0,0)
        self.sem2.layout.addWidget(self.title_label2,0,1)
        self.sem2.layout.addWidget(self.title_label3,0,2)

        self.sem2.layout.addWidget(self.sem2_sub1_credit_label,1,0)
        self.sem2.layout.addWidget(self.sem2_sub1_name_label,1,1)
        self.sem2.layout.addWidget(self.sem2_sub1_grade_label,1,2)
        
        self.sem2.layout.addWidget(self.sem2_sub2_credit_label,2,0)
        self.sem2.layout.addWidget(self.sem2_sub2_name_label,2,1)
        self.sem2.layout.addWidget(self.sem2_sub2_grade_label,2,2)
        
        self.sem2.layout.addWidget(self.sem2_sub3_credit_label,3,0)
        self.sem2.layout.addWidget(self.sem2_sub3_name_label,3,1)
        self.sem2.layout.addWidget(self.sem2_sub3_grade_label,3,2)

        self.sem2.layout.addWidget(self.sem2_sub4_credit_label,4,0)
        self.sem2.layout.addWidget(self.sem2_sub4_name_label,4,1)
        self.sem2.layout.addWidget(self.sem2_sub4_grade_label,4,2)

        self.sem2.layout.addWidget(self.sem2_sub5_credit_label,5,0)
        self.sem2.layout.addWidget(self.sem2_sub5_name_label,5,1)
        self.sem2.layout.addWidget(self.sem2_sub5_grade_label,5,2)

        self.sem2.layout.addWidget(self.sem2_sub6_credit_label,6,0)
        self.sem2.layout.addWidget(self.sem2_sub6_name_label,6,1)
        self.sem2.layout.addWidget(self.sem2_sub6_grade_label,6,2)

        self.sem2.layout.addWidget(self.sem2_sub7_credit_label,7,0)
        self.sem2.layout.addWidget(self.sem2_sub7_name_label,7,1)
        self.sem2.layout.addWidget(self.sem2_sub7_grade_label,7,2)

        self.sem2_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem2_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem2.layout.addWidget(self.sem2_gpa_label,8,0)
        self.sem2_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem2.layout.addWidget(self.sem2_gpa_result_label,8,2)

        self.sem2.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem2.setLayout(self.sem2.layout)


        # Semester 3
        self.sem3.layout = QtWidgets.QGridLayout()
        self.sem3.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem3_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem3_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub1_name_label = QtWidgets.QLabel('Linear Algebra')
        self.sem3_sub1_grade_label = QtWidgets.QLabel()

        self.sem3_sub2_credit_label = QtWidgets.QLabel('4')
        self.sem3_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub2_name_label = QtWidgets.QLabel('Electronic Circuits I')
        self.sem3_sub2_grade_label = QtWidgets.QLabel()

        self.sem3_sub3_credit_label = QtWidgets.QLabel('4')
        self.sem3_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub3_name_label = QtWidgets.QLabel('Signals & System')
        self.sem3_sub3_grade_label = QtWidgets.QLabel()

        self.sem3_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem3_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub4_name_label = QtWidgets.QLabel('Electromagnetic Fields & Waves ')
        self.sem3_sub4_grade_label = QtWidgets.QLabel()

        self.sem3_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem3_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub5_name_label = QtWidgets.QLabel('Digital System Design')
        self.sem3_sub5_grade_label = QtWidgets.QLabel()
        
        self.sem3_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem3_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub6_name_label = QtWidgets.QLabel('Electronic Design Lab')
        self.sem3_sub6_grade_label = QtWidgets.QLabel()

        self.sem3_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem3_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3_sub7_name_label = QtWidgets.QLabel('Electrical & Measurements Lab')
        self.sem3_sub7_grade_label = QtWidgets.QLabel()

        self.sem3.layout.addWidget(self.title_label1,0,0)
        self.sem3.layout.addWidget(self.title_label2,0,1)
        self.sem3.layout.addWidget(self.title_label3,0,2)

        self.sem3.layout.addWidget(self.sem3_sub1_credit_label,1,0)
        self.sem3.layout.addWidget(self.sem3_sub1_name_label,1,1)
        self.sem3.layout.addWidget(self.sem3_sub1_grade_label,1,2)
        
        self.sem3.layout.addWidget(self.sem3_sub2_credit_label,2,0)
        self.sem3.layout.addWidget(self.sem3_sub2_name_label,2,1)
        self.sem3.layout.addWidget(self.sem3_sub2_grade_label,2,2)
        
        self.sem3.layout.addWidget(self.sem3_sub3_credit_label,3,0)
        self.sem3.layout.addWidget(self.sem3_sub3_name_label,3,1)
        self.sem3.layout.addWidget(self.sem3_sub3_grade_label,3,2)

        self.sem3.layout.addWidget(self.sem3_sub4_credit_label,4,0)
        self.sem3.layout.addWidget(self.sem3_sub4_name_label,4,1)
        self.sem3.layout.addWidget(self.sem3_sub4_grade_label,4,2)

        self.sem3.layout.addWidget(self.sem3_sub5_credit_label,5,0)
        self.sem3.layout.addWidget(self.sem3_sub5_name_label,5,1)
        self.sem3.layout.addWidget(self.sem3_sub5_grade_label,5,2)

        self.sem3.layout.addWidget(self.sem3_sub6_credit_label,6,0)
        self.sem3.layout.addWidget(self.sem3_sub6_name_label,6,1)
        self.sem3.layout.addWidget(self.sem3_sub6_grade_label,6,2)

        self.sem3.layout.addWidget(self.sem3_sub7_credit_label,7,0)
        self.sem3.layout.addWidget(self.sem3_sub7_name_label,7,1)
        self.sem3.layout.addWidget(self.sem3_sub7_grade_label,7,2)

        self.sem3_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem3_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem3.layout.addWidget(self.sem3_gpa_label,8,0)
        self.sem3_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem3.layout.addWidget(self.sem3_gpa_result_label,8,2)

        self.sem3.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem3.setLayout(self.sem3.layout)


        # Semester 4
        self.sem4.layout = QtWidgets.QGridLayout()
        self.sem4.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem4_sub1_credit_label = QtWidgets.QLabel('4')
        self.sem4_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub1_name_label = QtWidgets.QLabel('Electronic Circuits II')
        self.sem4_sub1_grade_label = QtWidgets.QLabel()

        self.sem4_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub2_name_label = QtWidgets.QLabel('Transmission Line & Wave Guides')
        self.sem4_sub2_grade_label = QtWidgets.QLabel()

        self.sem4_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub3_name_label = QtWidgets.QLabel('Communication Theory')
        self.sem4_sub3_grade_label = QtWidgets.QLabel()

        self.sem4_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub4_name_label = QtWidgets.QLabel('Digital Signal Processing')
        self.sem4_sub4_grade_label = QtWidgets.QLabel()

        self.sem4_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub5_name_label = QtWidgets.QLabel('Linear Intergrated Circuits')
        self.sem4_sub5_grade_label = QtWidgets.QLabel()
        
        self.sem4_sub6_credit_label = QtWidgets.QLabel('3')
        self.sem4_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub6_name_label = QtWidgets.QLabel('Environmental Science')
        self.sem4_sub6_grade_label = QtWidgets.QLabel()

        self.sem4_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem4_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub7_name_label = QtWidgets.QLabel('Digital Signal Lab')
        self.sem4_sub7_grade_label = QtWidgets.QLabel()

        self.sem4_sub8_credit_label = QtWidgets.QLabel('2')
        self.sem4_sub8_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4_sub8_name_label = QtWidgets.QLabel('Integrated Circuits Lab')
        self.sem4_sub8_grade_label = QtWidgets.QLabel()

        self.sem4.layout.addWidget(self.title_label1,0,0)
        self.sem4.layout.addWidget(self.title_label2,0,1)
        self.sem4.layout.addWidget(self.title_label3,0,2)

        self.sem4.layout.addWidget(self.sem4_sub1_credit_label,1,0)
        self.sem4.layout.addWidget(self.sem4_sub1_name_label,1,1)
        self.sem4.layout.addWidget(self.sem4_sub1_grade_label,1,2)
        
        self.sem4.layout.addWidget(self.sem4_sub2_credit_label,2,0)
        self.sem4.layout.addWidget(self.sem4_sub2_name_label,2,1)
        self.sem4.layout.addWidget(self.sem4_sub2_grade_label,2,2)
        
        self.sem4.layout.addWidget(self.sem4_sub3_credit_label,3,0)
        self.sem4.layout.addWidget(self.sem4_sub3_name_label,3,1)
        self.sem4.layout.addWidget(self.sem4_sub3_grade_label,3,2)

        self.sem4.layout.addWidget(self.sem4_sub4_credit_label,4,0)
        self.sem4.layout.addWidget(self.sem4_sub4_name_label,4,1)
        self.sem4.layout.addWidget(self.sem4_sub4_grade_label,4,2)

        self.sem4.layout.addWidget(self.sem4_sub5_credit_label,5,0)
        self.sem4.layout.addWidget(self.sem4_sub5_name_label,5,1)
        self.sem4.layout.addWidget(self.sem4_sub5_grade_label,5,2)

        self.sem4.layout.addWidget(self.sem4_sub6_credit_label,6,0)
        self.sem4.layout.addWidget(self.sem4_sub6_name_label,6,1)
        self.sem4.layout.addWidget(self.sem4_sub6_grade_label,6,2)

        self.sem4.layout.addWidget(self.sem4_sub7_credit_label,7,0)
        self.sem4.layout.addWidget(self.sem4_sub7_name_label,7,1)
        self.sem4.layout.addWidget(self.sem4_sub7_grade_label,7,2)

        self.sem4.layout.addWidget(self.sem4_sub8_credit_label,8,0)
        self.sem4.layout.addWidget(self.sem4_sub8_name_label,8,1)
        self.sem4.layout.addWidget(self.sem4_sub8_grade_label,8,2)

        self.sem4_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem4_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem4.layout.addWidget(self.sem4_gpa_label,9,0)
        self.sem4_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem4.layout.addWidget(self.sem4_gpa_result_label,9,2)

        self.sem4.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem4.setLayout(self.sem4.layout)

        # Semester 5
        self.sem5.layout = QtWidgets.QGridLayout()
        self.sem5.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem5_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub1_name_label = QtWidgets.QLabel('Antennas')
        self.sem5_sub1_grade_label = QtWidgets.QLabel()

        self.sem5_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub2_name_label = QtWidgets.QLabel('Digital Communication')
        self.sem5_sub2_grade_label = QtWidgets.QLabel()

        self.sem5_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub3_name_label = QtWidgets.QLabel('Microprocessors')
        self.sem5_sub3_grade_label = QtWidgets.QLabel()

        self.sem5_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub4_name_label = QtWidgets.QLabel('Control Systems')
        self.sem5_sub4_grade_label = QtWidgets.QLabel()

        self.sem5_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub5_name_label = QtWidgets.QLabel('Principles of Management       ')
        self.sem5_sub5_grade_label = QtWidgets.QLabel()
        
        self.sem5_sub6_credit_label = QtWidgets.QLabel('3')
        self.sem5_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub6_name_label = QtWidgets.QLabel('Professinal Elective 1')
        self.sem5_sub6_grade_label = QtWidgets.QLabel()

        self.sem5_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem5_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub7_name_label = QtWidgets.QLabel('Microprocessors Lab')
        self.sem5_sub7_grade_label = QtWidgets.QLabel()

        self.sem5_sub8_credit_label = QtWidgets.QLabel('2')
        self.sem5_sub8_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub8_name_label = QtWidgets.QLabel('Digital Communication Lab')
        self.sem5_sub8_grade_label = QtWidgets.QLabel()

        self.sem5_sub9_credit_label = QtWidgets.QLabel('2')
        self.sem5_sub9_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5_sub9_name_label = QtWidgets.QLabel('Summer Project')
        self.sem5_sub9_grade_label = QtWidgets.QLabel()

        self.sem5.layout.addWidget(self.title_label1,0,0)
        self.sem5.layout.addWidget(self.title_label2,0,1)
        self.sem5.layout.addWidget(self.title_label3,0,2)

        self.sem5.layout.addWidget(self.sem5_sub1_credit_label,1,0)
        self.sem5.layout.addWidget(self.sem5_sub1_name_label,1,1)
        self.sem5.layout.addWidget(self.sem5_sub1_grade_label,1,2)

        self.sem5.layout.addWidget(self.sem5_sub2_credit_label,2,0)
        self.sem5.layout.addWidget(self.sem5_sub2_name_label,2,1)
        self.sem5.layout.addWidget(self.sem5_sub2_grade_label,2,2)

        self.sem5.layout.addWidget(self.sem5_sub3_credit_label,3,0)
        self.sem5.layout.addWidget(self.sem5_sub3_name_label,3,1)
        self.sem5.layout.addWidget(self.sem5_sub3_grade_label,3,2)
        
        self.sem5.layout.addWidget(self.sem5_sub4_credit_label,4,0)
        self.sem5.layout.addWidget(self.sem5_sub4_name_label,4,1)
        self.sem5.layout.addWidget(self.sem5_sub4_grade_label,4,2)
        
        self.sem5.layout.addWidget(self.sem5_sub5_credit_label,5,0)
        self.sem5.layout.addWidget(self.sem5_sub5_name_label,5,1)
        self.sem5.layout.addWidget(self.sem5_sub5_grade_label,5,2)
        
        self.sem5.layout.addWidget(self.sem5_sub6_credit_label,6,0)
        self.sem5.layout.addWidget(self.sem5_sub6_name_label,6,1)
        self.sem5.layout.addWidget(self.sem5_sub6_grade_label,6,2)
        
        self.sem5.layout.addWidget(self.sem5_sub7_credit_label,7,0)
        self.sem5.layout.addWidget(self.sem5_sub7_name_label,7,1)
        self.sem5.layout.addWidget(self.sem5_sub7_grade_label,7,2)
        
        self.sem5.layout.addWidget(self.sem5_sub8_credit_label,8,0)
        self.sem5.layout.addWidget(self.sem5_sub8_name_label,8,1)
        self.sem5.layout.addWidget(self.sem5_sub8_grade_label,8,2)
        
        self.sem5.layout.addWidget(self.sem5_sub9_credit_label,9,0)
        self.sem5.layout.addWidget(self.sem5_sub9_name_label,9,1)
        self.sem5.layout.addWidget(self.sem5_sub9_grade_label,9,2)

        self.sem5_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem5_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem5.layout.addWidget(self.sem5_gpa_label,10,0)
        self.sem5_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem5.layout.addWidget(self.sem5_gpa_result_label,10,2)

        self.sem5.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem5.setLayout(self.sem5.layout)

        # Semester 6
        self.sem6.layout = QtWidgets.QGridLayout()
        self.sem6.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem6_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub1_name_label = QtWidgets.QLabel('Digital VLSI')
        self.sem6_sub1_grade_label = QtWidgets.QLabel()

        self.sem6_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub2_name_label = QtWidgets.QLabel('Wireless Communications        ')
        self.sem6_sub2_grade_label = QtWidgets.QLabel()

        self.sem6_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub3_name_label = QtWidgets.QLabel('Communication Networks')
        self.sem6_sub3_grade_label = QtWidgets.QLabel()

        self.sem6_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub4_name_label = QtWidgets.QLabel('Professional Elective II')
        self.sem6_sub4_grade_label = QtWidgets.QLabel()

        self.sem6_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub5_name_label = QtWidgets.QLabel('Open Elective I')
        self.sem6_sub5_grade_label = QtWidgets.QLabel()
        
        self.sem6_sub6_credit_label = QtWidgets.QLabel('3')
        self.sem6_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub6_name_label = QtWidgets.QLabel('Values & Ethics')
        self.sem6_sub6_grade_label = QtWidgets.QLabel()

        self.sem6_sub7_credit_label = QtWidgets.QLabel('2')
        self.sem6_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub7_name_label = QtWidgets.QLabel('VLSI Lab')
        self.sem6_sub7_grade_label = QtWidgets.QLabel()

        self.sem6_sub8_credit_label = QtWidgets.QLabel('2')
        self.sem6_sub8_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6_sub8_name_label = QtWidgets.QLabel('Wireless Communications Lab')
        self.sem6_sub8_grade_label = QtWidgets.QLabel()

        self.sem6.layout.addWidget(self.title_label1,0,0)
        self.sem6.layout.addWidget(self.title_label2,0,1)
        self.sem6.layout.addWidget(self.title_label3,0,2)

        self.sem6.layout.addWidget(self.sem6_sub1_credit_label,1,0)
        self.sem6.layout.addWidget(self.sem6_sub1_name_label,1,1)
        self.sem6.layout.addWidget(self.sem6_sub1_grade_label,1,2)

        self.sem6.layout.addWidget(self.sem6_sub2_credit_label,2,0)
        self.sem6.layout.addWidget(self.sem6_sub2_name_label,2,1)
        self.sem6.layout.addWidget(self.sem6_sub2_grade_label,2,2)

        self.sem6.layout.addWidget(self.sem6_sub3_credit_label,3,0)
        self.sem6.layout.addWidget(self.sem6_sub3_name_label,3,1)
        self.sem6.layout.addWidget(self.sem6_sub3_grade_label,3,2)

        self.sem6.layout.addWidget(self.sem6_sub4_credit_label,4,0)
        self.sem6.layout.addWidget(self.sem6_sub4_name_label,4,1)
        self.sem6.layout.addWidget(self.sem6_sub4_grade_label,4,2)

        self.sem6.layout.addWidget(self.sem6_sub5_credit_label,5,0)
        self.sem6.layout.addWidget(self.sem6_sub5_name_label,5,1)
        self.sem6.layout.addWidget(self.sem6_sub5_grade_label,5,2)

        self.sem6.layout.addWidget(self.sem6_sub6_credit_label,6,0)
        self.sem6.layout.addWidget(self.sem6_sub6_name_label,6,1)
        self.sem6.layout.addWidget(self.sem6_sub6_grade_label,6,2)

        self.sem6.layout.addWidget(self.sem6_sub7_credit_label,7,0)
        self.sem6.layout.addWidget(self.sem6_sub7_name_label,7,1)
        self.sem6.layout.addWidget(self.sem6_sub7_grade_label,7,2)

        self.sem6.layout.addWidget(self.sem6_sub8_credit_label,8,0)
        self.sem6.layout.addWidget(self.sem6_sub8_name_label,8,1)
        self.sem6.layout.addWidget(self.sem6_sub8_grade_label,8,2)

        self.sem6_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem6_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem6.layout.addWidget(self.sem6_gpa_label,9,0)
        self.sem6_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem6.layout.addWidget(self.sem6_gpa_result_label,9,2)

        self.sem6.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem6.setLayout(self.sem6.layout)

        # Semester 7
        self.sem7.layout = QtWidgets.QGridLayout()
        self.sem7.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem7_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub1_name_label = QtWidgets.QLabel('Optical Wave Communications    ')
        self.sem7_sub1_grade_label = QtWidgets.QLabel()

        self.sem7_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub2_name_label = QtWidgets.QLabel('Human Relations')
        self.sem7_sub2_grade_label = QtWidgets.QLabel()

        self.sem7_sub3_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub3_name_label = QtWidgets.QLabel('Professional Elective III')
        self.sem7_sub3_grade_label = QtWidgets.QLabel()

        self.sem7_sub4_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub4_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub4_name_label = QtWidgets.QLabel('Professional Elective IV')
        self.sem7_sub4_grade_label = QtWidgets.QLabel()

        self.sem7_sub5_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub5_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub5_name_label = QtWidgets.QLabel('Open Elective II')
        self.sem7_sub5_grade_label = QtWidgets.QLabel()
        
        self.sem7_sub6_credit_label = QtWidgets.QLabel('2')
        self.sem7_sub6_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub6_name_label = QtWidgets.QLabel('High Frequency Lab')
        self.sem7_sub6_grade_label = QtWidgets.QLabel()

        self.sem7_sub7_credit_label = QtWidgets.QLabel('3')
        self.sem7_sub7_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7_sub7_name_label = QtWidgets.QLabel('Project I')
        self.sem7_sub7_grade_label = QtWidgets.QLabel()

        self.sem7.layout.addWidget(self.title_label1,0,0)
        self.sem7.layout.addWidget(self.title_label2,0,1)
        self.sem7.layout.addWidget(self.title_label3,0,2)

        self.sem7.layout.addWidget(self.sem7_sub1_credit_label,1,0)
        self.sem7.layout.addWidget(self.sem7_sub1_name_label,1,1)
        self.sem7.layout.addWidget(self.sem7_sub1_grade_label,1,2)
        
        self.sem7.layout.addWidget(self.sem7_sub2_credit_label,2,0)
        self.sem7.layout.addWidget(self.sem7_sub2_name_label,2,1)
        self.sem7.layout.addWidget(self.sem7_sub2_grade_label,2,2)
        
        self.sem7.layout.addWidget(self.sem7_sub3_credit_label,3,0)
        self.sem7.layout.addWidget(self.sem7_sub3_name_label,3,1)
        self.sem7.layout.addWidget(self.sem7_sub3_grade_label,3,2)

        self.sem7.layout.addWidget(self.sem7_sub4_credit_label,4,0)
        self.sem7.layout.addWidget(self.sem7_sub4_name_label,4,1)
        self.sem7.layout.addWidget(self.sem7_sub4_grade_label,4,2)

        self.sem7.layout.addWidget(self.sem7_sub5_credit_label,5,0)
        self.sem7.layout.addWidget(self.sem7_sub5_name_label,5,1)
        self.sem7.layout.addWidget(self.sem7_sub5_grade_label,5,2)

        self.sem7.layout.addWidget(self.sem7_sub6_credit_label,6,0)
        self.sem7.layout.addWidget(self.sem7_sub6_name_label,6,1)
        self.sem7.layout.addWidget(self.sem7_sub6_grade_label,6,2)

        self.sem7.layout.addWidget(self.sem7_sub7_credit_label,7,0)
        self.sem7.layout.addWidget(self.sem7_sub7_name_label,7,1)
        self.sem7.layout.addWidget(self.sem7_sub7_grade_label,7,2)

        self.sem7_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem7_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem7.layout.addWidget(self.sem7_gpa_label,8,0)
        self.sem7_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem7.layout.addWidget(self.sem7_gpa_result_label,8,2)

        self.sem7.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem7.setLayout(self.sem7.layout)

        # Semester 8
        self.sem8.layout = QtWidgets.QGridLayout()
        self.sem8.layout.setSpacing(10)

        self.title_label1 = QtWidgets.QLabel('Credit')
        self.title_label1.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label2 = QtWidgets.QLabel('Subject')
        self.title_label3 = QtWidgets.QLabel('Grade')

        self.sem8_sub1_credit_label = QtWidgets.QLabel('3')
        self.sem8_sub1_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem8_sub1_name_label = QtWidgets.QLabel('Professional Elective V')
        self.sem8_sub1_grade_label = QtWidgets.QLabel()

        self.sem8_sub2_credit_label = QtWidgets.QLabel('3')
        self.sem8_sub2_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem8_sub2_name_label = QtWidgets.QLabel('Professional Elective VI       ')
        self.sem8_sub2_grade_label = QtWidgets.QLabel()

        self.sem8_sub3_credit_label = QtWidgets.QLabel('8')
        self.sem8_sub3_credit_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem8_sub3_name_label = QtWidgets.QLabel('Project II')
        self.sem8_sub3_grade_label = QtWidgets.QLabel()

        
        self.sem8.layout.addWidget(self.title_label1,0,0)
        self.sem8.layout.addWidget(self.title_label2,0,1)
        self.sem8.layout.addWidget(self.title_label3,0,2)

        self.sem8.layout.addWidget(self.sem8_sub1_credit_label,1,0)
        self.sem8.layout.addWidget(self.sem8_sub1_name_label,1,1)
        self.sem8.layout.addWidget(self.sem8_sub1_grade_label,1,2)
        
        self.sem8.layout.addWidget(self.sem8_sub2_credit_label,2,0)
        self.sem8.layout.addWidget(self.sem8_sub2_name_label,2,1)
        self.sem8.layout.addWidget(self.sem8_sub2_grade_label,2,2)
        
        self.sem8.layout.addWidget(self.sem8_sub3_credit_label,3,0)
        self.sem8.layout.addWidget(self.sem8_sub3_name_label,3,1)
        self.sem8.layout.addWidget(self.sem8_sub3_grade_label,3,2)

        self.sem8.layout.setAlignment(QtCore.Qt.AlignTop)
        self.sem8.setLayout(self.sem8.layout)

        self.sem8_gpa_label = QtWidgets.QLabel('GPA:')
        self.sem8_gpa_result_label = QtWidgets.QLabel('0.00')

        self.sem8.layout.addWidget(self.sem8_gpa_label,4,0)
        self.sem8_gpa_label.setAlignment(QtCore.Qt.AlignCenter)
        self.sem8.layout.addWidget(self.sem8_gpa_result_label,4,2)


        self.hbox_tab = QtWidgets.QHBoxLayout()
        self.hbox_tab.addWidget(self.tabs)

        ######################################################
        self.cancel_button = QtWidgets.QPushButton(self)
        self.clear_button = QtWidgets.QPushButton(self)
        self.search_button = QtWidgets.QPushButton(self)

        self.cancel_button.setText("Cancel")
        self.clear_button.setText("Clear")
        self.search_button.setText("Search")

        self.clear_button.clicked.connect(self.clear_data)
        self.cancel_button.clicked.connect(self.cancel_data)
        self.search_button.clicked.connect(self.search_data)

        self.hbox = QtWidgets.QHBoxLayout()
        self.hbox.addStretch(1)

        self.hbox.addWidget(self.search_button)
        self.hbox.addWidget(self.clear_button) 
        self.hbox.addWidget(self.cancel_button) 
        ######################################################

        self.vbox = QtWidgets.QVBoxLayout()
        self.vbox.addLayout(self.grid)
        self.vbox.addLayout(self.hbox_tab)
        self.vbox.addLayout(self.hbox)

        self.setLayout(self.vbox)
    
    

    def search_data(self):
        print('searching ',self.roll_edit.text())
        roll = self.roll_edit.text()
        department=[" ",
                    "CSE",
                    "IT",
                    "ECE",
                    "BIO MEDICAL",
                    "EEE",
                    "PRINTING",
                    "MINING",
                    "MANUFACTURING",
                    "INDUSTRIAL",
                    "MECHANICAL",
                    "MATERIAL SCIENCE",
                    "CIVIL",
                    "GEO INFORMATICS"]

        usr_name = os.getlogin()
        wb = openpyxl.load_workbook("gpa.xlsx")
        sheet = wb.active

        available_flag = 0
        for i in range(2,sheet.max_row+1):
            cell_obj = sheet.cell(row=i,column=2)
            if roll == cell_obj.value:
                available_flag = 1
        
        if available_flag == 1:
            for i in range(2,sheet.max_row+1):
                cell_obj = sheet.cell(row=i,column=2)
                if(roll == cell_obj.value):
                    self.name_edit.setText((sheet.cell(row=i,column=1)).value)
                    self.phno_edit.setText((sheet.cell(row=i,column=3)).value)
                    self.mail_edit.setText((sheet.cell(row=i,column=4)).value)
                    self.dept_edit.setText((sheet.cell(row=i,column=5)).value)
                    self.sem1_sub1_grade_label.setText((sheet.cell(row=i,column=6)).value)
                    self.sem1_sub2_grade_label.setText((sheet.cell(row=i,column=7)).value)
                    self.sem1_sub3_grade_label.setText((sheet.cell(row=i,column=8)).value)
                    self.sem1_sub4_grade_label.setText((sheet.cell(row=i,column=9)).value)
                    self.sem1_sub5_grade_label.setText((sheet.cell(row=i,column=10)).value)
                    self.sem1_sub6_grade_label.setText((sheet.cell(row=i,column=11)).value)
                    self.sem1_sub7_grade_label.setText((sheet.cell(row=i,column=12)).value)
                    self.sem2_sub1_grade_label.setText((sheet.cell(row=i,column=13)).value)
                    self.sem2_sub2_grade_label.setText((sheet.cell(row=i,column=14)).value)
                    self.sem2_sub3_grade_label.setText((sheet.cell(row=i,column=15)).value)
                    self.sem2_sub4_grade_label.setText((sheet.cell(row=i,column=16)).value)
                    self.sem2_sub5_grade_label.setText((sheet.cell(row=i,column=17)).value)
                    self.sem2_sub6_grade_label.setText((sheet.cell(row=i,column=18)).value)
                    self.sem2_sub7_grade_label.setText((sheet.cell(row=i,column=19)).value)
                    self.sem3_sub1_grade_label.setText((sheet.cell(row=i,column=20)).value)
                    self.sem3_sub2_grade_label.setText((sheet.cell(row=i,column=21)).value)
                    self.sem3_sub3_grade_label.setText((sheet.cell(row=i,column=22)).value)
                    self.sem3_sub4_grade_label.setText((sheet.cell(row=i,column=23)).value)
                    self.sem3_sub5_grade_label.setText((sheet.cell(row=i,column=24)).value)
                    self.sem3_sub6_grade_label.setText((sheet.cell(row=i,column=25)).value)
                    self.sem3_sub7_grade_label.setText((sheet.cell(row=i,column=26)).value)
                    self.sem4_sub1_grade_label.setText((sheet.cell(row=i,column=27)).value)
                    self.sem4_sub2_grade_label.setText((sheet.cell(row=i,column=28)).value)
                    self.sem4_sub3_grade_label.setText((sheet.cell(row=i,column=29)).value)
                    self.sem4_sub4_grade_label.setText((sheet.cell(row=i,column=30)).value)
                    self.sem4_sub5_grade_label.setText((sheet.cell(row=i,column=31)).value)
                    self.sem4_sub6_grade_label.setText((sheet.cell(row=i,column=32)).value)
                    self.sem4_sub7_grade_label.setText((sheet.cell(row=i,column=33)).value)
                    self.sem4_sub8_grade_label.setText((sheet.cell(row=i,column=34)).value)
                    self.sem5_sub1_grade_label.setText((sheet.cell(row=i,column=35)).value)
                    self.sem5_sub2_grade_label.setText((sheet.cell(row=i,column=36)).value)
                    self.sem5_sub3_grade_label.setText((sheet.cell(row=i,column=37)).value)
                    self.sem5_sub4_grade_label.setText((sheet.cell(row=i,column=38)).value)
                    self.sem5_sub5_grade_label.setText((sheet.cell(row=i,column=39)).value)
                    self.sem5_sub6_grade_label.setText((sheet.cell(row=i,column=40)).value)
                    self.sem5_sub7_grade_label.setText((sheet.cell(row=i,column=41)).value)
                    self.sem5_sub8_grade_label.setText((sheet.cell(row=i,column=42)).value)
                    self.sem5_sub9_grade_label.setText((sheet.cell(row=i,column=43)).value)
                    self.sem6_sub1_grade_label.setText((sheet.cell(row=i,column=44)).value)
                    self.sem6_sub2_grade_label.setText((sheet.cell(row=i,column=45)).value)
                    self.sem6_sub3_grade_label.setText((sheet.cell(row=i,column=46)).value)
                    self.sem6_sub4_grade_label.setText((sheet.cell(row=i,column=47)).value)
                    self.sem6_sub5_grade_label.setText((sheet.cell(row=i,column=48)).value)
                    self.sem6_sub6_grade_label.setText((sheet.cell(row=i,column=49)).value)
                    self.sem6_sub7_grade_label.setText((sheet.cell(row=i,column=50)).value)
                    self.sem6_sub8_grade_label.setText((sheet.cell(row=i,column=51)).value)
                    self.sem7_sub1_grade_label.setText((sheet.cell(row=i,column=52)).value)
                    self.sem7_sub2_grade_label.setText((sheet.cell(row=i,column=53)).value)
                    self.sem7_sub3_grade_label.setText((sheet.cell(row=i,column=54)).value)
                    self.sem7_sub4_grade_label.setText((sheet.cell(row=i,column=55)).value)
                    self.sem7_sub5_grade_label.setText((sheet.cell(row=i,column=56)).value)
                    self.sem7_sub6_grade_label.setText((sheet.cell(row=i,column=57)).value)
                    self.sem7_sub7_grade_label.setText((sheet.cell(row=i,column=58)).value)
                    self.sem8_sub1_grade_label.setText((sheet.cell(row=i,column=59)).value)
                    self.sem8_sub2_grade_label.setText((sheet.cell(row=i,column=60)).value)
                    self.sem8_sub3_grade_label.setText((sheet.cell(row=i,column=61)).value)

                    # gpa
                    self.sem1_gpa_result_label.setText(str((sheet.cell(row=i,column=62)).value))
                    self.sem2_gpa_result_label.setText(str((sheet.cell(row=i,column=63)).value))
                    self.sem3_gpa_result_label.setText(str((sheet.cell(row=i,column=64)).value))
                    self.sem4_gpa_result_label.setText(str((sheet.cell(row=i,column=65)).value))
                    self.sem5_gpa_result_label.setText(str((sheet.cell(row=i,column=66)).value))
                    self.sem6_gpa_result_label.setText(str((sheet.cell(row=i,column=67)).value))
                    self.sem7_gpa_result_label.setText(str((sheet.cell(row=i,column=68)).value))
                    self.sem8_gpa_result_label.setText(str((sheet.cell(row=i,column=69)).value))

                    # cgpa
                    self.cgpa_edit.setText(str((sheet.cell(row=i,column=70)).value))
                    
                    break
        else:
            msg = QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setWindowTitle('Warning')
            msg.setText('Roll Number is Invalid!')
            msg.setInformativeText('Kindly register or enter a valid roll number.')
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            retval = msg.exec_()

    def clear_data(self):
        print('clear')
        self.roll_edit.clear()

    def cancel_data(self):
        print('cancel')
        self.close()


class UI_MainWindow(QtWidgets.QWidget):

    def setupUi(self, MainWindow):

        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(600,500)

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setStyleSheet("\n"
            "background:url(resources/images/background2.png);\n"
            "color:white;\n"
            "")
        # image label
        self.img_label1 = QtWidgets.QLabel(self.centralwidget)
        self.img_label1.setGeometry(QtCore.QRect(15,15,100,100))
        self.img_label1.setText("")
        self.img_label1.setPixmap(QtGui.QPixmap("resources/images/ceg_logo_bw.jpg"))
        self.img_label1.setScaledContents(True)
        self.img_label1.setObjectName("img_label1")

        # label1 - heading
        self.label_1 = QtWidgets.QLabel(self.centralwidget)
        self.label_1.setGeometry(QtCore.QRect(120, 15, 460, 100))
        font = QtGui.QFont("Libeartion serif",16)
        self.label_1.setFont(font)
        self.label_1.setStyleSheet(
            "background:rgba(200,200,200,20);\n"
            "border-radius : 5px;\n"
            "")
        self.label_1.setAlignment(QtCore.Qt.AlignCenter)
        self.label_1.setObjectName("label_1")

        # label2 - new contact
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(15, 150, 500, 30))
        font = QtGui.QFont("Libeartion serif",11)
        self.label_2.setFont(font)
        self.label_2.setStyleSheet(
            "background:rgba(200,200,200,80);\n"
            "border-radius : 5px;\n"
            "")
        self.label_2.setObjectName("label_2")

        # button1 - new contact
        self.centralwidget.setObjectName("centralwidget")
        self.newButton = QtWidgets.QPushButton(self.centralwidget)
        self.newButton.setGeometry(QtCore.QRect(420, 150, 150, 30))
        font = QtGui.QFont('Sans-serif', 12)
        self.newButton.setFont(font)
        self.newButton.setStyleSheet(
            "QPushButton:hover{\n"
            "background:rgb(15,15,15);\n"
            "\n"
            "}")
        self.newButton.setObjectName("newButton")

        # label3 - update contact
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(15, 200, 500, 30))
        font = QtGui.QFont("Libeartion serif",11)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet(
            "background:rgba(200,200,200,80);\n"
            "border-radius : 5px;\n"
            "")
        self.label_3.setObjectName("label_3")

        # button2 - update contact
        self.centralwidget.setObjectName("centralwidget")
        self.upButton = QtWidgets.QPushButton(self.centralwidget)
        self.upButton.setGeometry(QtCore.QRect(420, 200, 150, 30))
        font = QtGui.QFont('Sans-serif', 12)
        self.upButton.setFont(font)
        self.upButton.setStyleSheet(
            "QPushButton:hover{\n"
            "background:rgb(15,15,15);\n"
            "\n"
            "}")
        self.upButton.setObjectName("upButton")

        # label4 - update contact
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(15, 250, 500, 30))
        font = QtGui.QFont("Libeartion serif",11)
        self.label_4.setFont(font)
        self.label_4.setStyleSheet(
            "background:rgba(200,200,200,80);\n"
            "border-radius : 5px;\n"
            "")
        self.label_4.setObjectName("label_3")

        # button3 - view contact
        self.centralwidget.setObjectName("centralwidget")
        self.viewButton = QtWidgets.QPushButton(self.centralwidget)
        self.viewButton.setGeometry(QtCore.QRect(420, 250, 150, 30))
        font = QtGui.QFont('Sans-serif', 12)
        self.viewButton.setFont(font)
        self.viewButton.setStyleSheet(
            "QPushButton:hover{\n"
            "background:rgb(15,15,15);\n"
            "\n"
            "}")
        self.viewButton.setObjectName("viewButton")

        ### Main Window
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)


    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "CGPA Calculator (R2019)"))
        self.label_1.setText(_translate("MainWindow", "College of Engineering Guindy\nAnna University"))
        self.label_2.setText(_translate("MainWindow", "  To register for a new student"))
        self.label_3.setText(_translate("MainWindow", "  To update details of a student"))
        self.label_4.setText(_translate("MainWindow", "  To view details of a student"))
        self.newButton.setText(_translate("MainWindow", "New"))
        self.newButton.clicked.connect(self.newUser)
        self.upButton.setText(_translate("MainWindow", "Update"))
        self.upButton.clicked.connect(self.updateUser)
        self.viewButton.setText(_translate("MainWindow", "View"))
        self.viewButton.clicked.connect(self.viewUser)

    
    def newUser(self):
        print("New User needs to be added!")
        self.ex1 = NewUser()
        self.ex1.show()

    def updateUser(self):
        print("Update User!")
        self.ex1 = UpdateUser()
        self.ex1.show()

    def viewUser(self):
        print("View User!")
        self.ex1 = ViewUser()
        self.ex1.show()

if __name__ == "__main__":

    # curr_dir = os.getcwd()
    #
    # usr_name = os.getlogin()
    # #os.chdir("/home/"+usr_name)
    #
    # cwd = os.getcwd()
    # for file in os.listdir(cwd):
    #     if not(os.path.isdir('GPA')):
    #         print('directory not exist')
    #         os.mkdir('GPA')
    #         break
    #
    # os.chdir("./GPA")
    cwd = os.getcwd()
    filename = 'gpa.xlsx'
    if filename not in list(os.listdir(cwd)):
        print('file doesnot exist')
        wb = Workbook()
        ws = wb.active

        ws.title = '2019-2023'
        ws.cell(row=1,column=1).value = 'name'
        ws.cell(row=1,column=2).value = 'roll'
        ws.cell(row=1,column=3).value = 'phone'
        ws.cell(row=1,column=4).value = 'mail'
        ws.cell(row=1,column=5).value = 'department'
        wb.save(filename='gpa.xlsx')

    # os.chdir(curr_dir)
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = UI_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
